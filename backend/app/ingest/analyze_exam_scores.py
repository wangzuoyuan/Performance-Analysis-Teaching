#!/usr/bin/env python3
"""Analyze exam score workbooks for class-teacher use (单学科化, 阶段7).

TotalScore 已退役：本 CLI/辅助分析管线不再包含任何总分（主三门/五门/九门/3+3）
相关的生成、参数、管线或 CSV/JSON 字段。只保留可工作的单科 score_long 与
subject average 分析（学籍1科目均分估算、班级总览、本班历史科目均分趋势）。
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from student_learning_html import build_learning_profile_outputs
except Exception:  # HTML output is helpful but should not block core analysis.
    build_learning_profile_outputs = None


SUBJECT_COLS = {
    "语文": (5, 6),
    "数学": (7, 8),
    "英语": (9, 10),
    "物理": (11, 12),
    "化学": (13, 14),
    "生物": (15, 16),
    "政治": (17, 18),
    "历史": (19, 20),
    "地理": (21, 22),
}

CLASS_AVG_SUBJECT_COLS = {
    "语文": 4,
    "数学": 5,
    "英语": 6,
    "物理": 7,
    "化学": 8,
    "生物": 9,
    "政治": 10,
    "历史": 11,
    "地理": 12,
}

EXAM_ORDER = {
    "高一第一学期12月月考": 1,
    "高一第一学期期末考试": 2,
    "高一第二学期期中考试": 3,
}

XUEJI_LABELS = {
    "1": "闵中学籍",
    "3": "文绮学籍",
    "4": "外省市体育生和复学学生",
}

SUBJECT_PCT_THRESHOLD = 0.10


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip().replace("\n", " ")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(value):
            return None
        return float(value)
    text = clean(value)
    if not text:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None


def intish(value: Any) -> int | None:
    value_num = num(value)
    if value_num is None:
        return None
    return int(round(value_num))


def is_numeric_class(value: Any) -> bool:
    text = clean(value)
    return bool(re.fullmatch(r"\d+", text))


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def canonical_exam_name(path: Path) -> str:
    stem = path.stem.strip()
    if "高一第二学期4月期中考试" in stem or "高一第二学期期中考试" in stem:
        return "高一第二学期期中考试"
    suffixes = [
        "主三门总分名次分段表",
        "主三门总分名次分段",
        "三门总分名次分段表",
        "三门总分名次分段",
        "五门总分名次分段表",
        "五门总分名次分段",
        "九门总分名次分段表",
        "九门总分名次分段",
        "班级均分表",
    ]
    name = stem
    for suffix in suffixes:
        if name.endswith(suffix):
            name = name[: -len(suffix)]
            break
    return name.strip()


def exam_sort_key(exam_name: str) -> tuple[int, str]:
    return (EXAM_ORDER.get(exam_name, 999), exam_name)


def row_values(ws: Any, row_idx: int, max_col: int | None = None) -> list[str]:
    max_col = max_col or ws.max_column
    return [clean(ws.cell(row_idx, col).value) for col in range(1, max_col + 1)]


def classify_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    row1 = row_values(ws, 1)
    row2 = row_values(ws, 2)
    row3 = row_values(ws, 3)
    workbook_type = "unknown"
    if len(row2) >= 4 and row2[:4] == ["学号", "班级", "学籍", "姓名"]:
        workbook_type = "student_scores"
    elif len(row1) >= 3 and row1[:3] == ["班级类型", "班级", "班主任"]:
        workbook_type = "class_averages"
    elif len(row2) >= 3 and row2[0] == "班级" and row2[1] == "班主任" and any("-" in x for x in row2[2:]):
        workbook_type = "rank_bands"
    title = row1[0] if row1 else ""
    return {
        "file": str(path),
        "file_name": path.name,
        "sha256": sha256_file(path),
        "exam": canonical_exam_name(path),
        "exam_order": EXAM_ORDER.get(canonical_exam_name(path)),
        "type": workbook_type,
        "sheets": wb.sheetnames,
        "active_sheet": ws.title,
        "rows": ws.max_row,
        "cols": ws.max_column,
        "title": title,
        "row1": row1,
        "row2": row2,
        "row3": row3,
        "warnings": [],
    }


def inspect_files(paths: list[Path]) -> dict[str, Any]:
    files = [classify_workbook(path) for path in paths]
    warnings: list[str] = []
    hashes: dict[str, list[str]] = defaultdict(list)
    by_exam_type: dict[tuple[str, str], list[str]] = defaultdict(list)
    for info in files:
        hashes[info["sha256"]].append(info["file_name"])
        by_exam_type[(info["exam"], info["type"])].append(info["file_name"])
        if info["type"] == "unknown":
            warnings.append(f"无法识别工作簿类型: {info['file_name']}")
    for digest, names in hashes.items():
        if len(names) > 1:
            warnings.append(f"发现内容完全相同的文件，疑似重复/错传: {', '.join(names)}")
    for (exam, workbook_type), names in by_exam_type.items():
        if workbook_type != "unknown" and len(names) > 1:
            warnings.append(f"同一考试和类型出现多个文件: {exam} / {workbook_type}: {', '.join(names)}")
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "files": files,
        "warnings": warnings,
    }


def parse_student_scores(path: Path, info: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """解析学生成绩明细表，返回 (students, score_long)。

    单学科化（阶段7）：不再返回 totals（TotalScore 已退役）。旧 Excel 含总分列
    仍可正常解析各单科成绩；总分列被忽略。
    """
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    score_long: list[dict[str, Any]] = []
    students: list[dict[str, Any]] = []
    exam = info["exam"]
    order = EXAM_ORDER.get(exam, 999)
    for row_idx in range(4, ws.max_row + 1):
        student_id = clean(ws.cell(row_idx, 1).value)
        if not student_id:
            continue
        class_name = clean(ws.cell(row_idx, 2).value)
        xueji = clean(ws.cell(row_idx, 3).value)
        name = clean(ws.cell(row_idx, 4).value)
        students.append({
            "exam": exam,
            "exam_order": order,
            "student_id": student_id,
            "class": class_name,
            "xueji": xueji,
            "xueji_label": XUEJI_LABELS.get(xueji, ""),
            "name": name,
            "source_file": path.name,
        })
        for subject, (score_col, pct_col) in SUBJECT_COLS.items():
            score = num(ws.cell(row_idx, score_col).value)
            pct = num(ws.cell(row_idx, pct_col).value)
            if score is None and pct is None:
                continue
            score_long.append({
                "exam": exam,
                "exam_order": order,
                "student_id": student_id,
                "class": class_name,
                "xueji": xueji,
                "name": name,
                "subject": subject,
                "score": score,
                "grade_percentile": pct,
                "source_file": path.name,
            })
    return students, score_long


def parse_class_averages(path: Path, info: dict[str, Any]) -> list[dict[str, Any]]:
    """解析班级均分表。单学科化（阶段7）：只解析单科均分列，不再解析
    主三门/五门/九门总分均分与名次（CLASS_AVG_TOTAL_COLS 已删除）。
    """
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    exam = info["exam"]
    order = EXAM_ORDER.get(exam, 999)
    current_class_type = ""
    rows: list[dict[str, Any]] = []
    for row_idx in range(3, ws.max_row + 1):
        if not any(ws.cell(row_idx, col).value not in (None, "") for col in range(1, ws.max_column + 1)):
            continue
        class_type_raw = clean(ws.cell(row_idx, 1).value)
        if class_type_raw:
            current_class_type = class_type_raw
        class_name = clean(ws.cell(row_idx, 2).value)
        teacher = clean(ws.cell(row_idx, 3).value)
        is_actual = is_numeric_class(class_name)
        row: dict[str, Any] = {
            "exam": exam,
            "exam_order": order,
            "class_type": current_class_type,
            "class": class_name,
            "teacher": teacher,
            "is_actual_class": is_actual,
            "source_file": path.name,
        }
        for subject, col in CLASS_AVG_SUBJECT_COLS.items():
            row[f"{subject}_avg"] = num(ws.cell(row_idx, col).value)
        rows.append(row)
    return rows


def parse_inputs(paths: list[Path]) -> dict[str, Any]:
    """解析输入文件集合。

    单学科化（阶段7）：不再解析 rank_bands（名次段表口径依赖主三门/五门/九门
    总分名次，§9 停止解析与输出）。旧名次段表文件可被识别但不产生跨学科包。
    """
    inspection = inspect_files(paths)
    students: list[dict[str, Any]] = []
    score_long: list[dict[str, Any]] = []
    class_averages: list[dict[str, Any]] = []
    warnings = list(inspection["warnings"])
    for info in inspection["files"]:
        path = Path(info["file"])
        if info["type"] == "student_scores":
            file_students, file_scores = parse_student_scores(path, info)
            students.extend(file_students)
            score_long.extend(file_scores)
        elif info["type"] == "class_averages":
            class_averages.extend(parse_class_averages(path, info))
        # rank_bands 不再解析（总分名次段口径已退役，§9）
    return {
        "inspection": inspection,
        "warnings": warnings,
        "students": students,
        "score_long": score_long,
        "class_averages": class_averages,
    }


def most_common_target_class(students: list[dict[str, Any]]) -> str:
    counts = Counter(row["class"] for row in students if row.get("class"))
    return counts.most_common(1)[0][0] if counts else "6"


def available_exams(rows: Iterable[dict[str, Any]]) -> list[str]:
    return sorted({row["exam"] for row in rows}, key=exam_sort_key)


def latest_exam(rows: Iterable[dict[str, Any]]) -> str | None:
    exams = available_exams(rows)
    return exams[-1] if exams else None


def previous_exam_name(exams: list[str], current: str | None) -> str | None:
    if current is None or current not in exams:
        return None
    idx = exams.index(current)
    if idx == 0:
        return None
    return exams[idx - 1]


def rows_by_key(rows: list[dict[str, Any]], *keys: str) -> dict[tuple[Any, ...], list[dict[str, Any]]]:
    grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[tuple(row.get(key) for key in keys)].append(row)
    return grouped


def latest_subject_percentiles(score_long: list[dict[str, Any]], student_id: str, exam: str) -> dict[str, float]:
    result: dict[str, float] = {}
    for row in score_long:
        if row["student_id"] == student_id and row["exam"] == exam and row.get("grade_percentile") is not None:
            result[row["subject"]] = float(row["grade_percentile"])
    return result


def previous_subject_percentiles(score_long: list[dict[str, Any]], student_id: str, subject: str, exam_order: int) -> tuple[str | None, float | None]:
    candidates = [
        row
        for row in score_long
        if row["student_id"] == student_id
        and row["subject"] == subject
        and row.get("grade_percentile") is not None
        and row.get("exam_order", 999) < exam_order
    ]
    if not candidates:
        return None, None
    candidates.sort(key=lambda row: row.get("exam_order", 999))
    row = candidates[-1]
    return row["exam"], float(row["grade_percentile"])


def recommended_subjects(
    score_long: list[dict[str, Any]],
    student_id: str,
    exam: str,
) -> tuple[list[str], list[str]]:
    """推荐关注科目：取百分位最弱的两科，附较上次退步幅度。

    单学科化（阶段7）：不再接收 total_percentile 参数（总分百分位已退役），
    偏科判断不再依赖「比总分百分位低」的口径。
    """
    pcts = latest_subject_percentiles(score_long, student_id, exam)
    if not pcts:
        return [], []
    weakest = sorted(pcts.items(), key=lambda item: item[1], reverse=True)[:2]
    recommended = [subject for subject, _ in weakest]
    reasons: list[str] = []
    for subject, pct in weakest:
        reason_bits = [f"{subject}百分位{pct:.3f}"]
        exam_order = EXAM_ORDER.get(exam, 999)
        prev_exam, prev_pct = previous_subject_percentiles(score_long, student_id, subject, exam_order)
        if prev_pct is not None and pct - prev_pct >= SUBJECT_PCT_THRESHOLD:
            reason_bits.append(f"较{prev_exam}退步{pct - prev_pct:.3f}")
        reasons.append("，".join(reason_bits))
    return recommended, reasons


def estimated_rank(value: float, comparators: Iterable[float | None]) -> int | None:
    usable = [float(item) for item in comparators if item not in (None, 0)]
    if not usable:
        return None
    return 1 + sum(1 for item in usable if item > value)


def build_xueji1_analysis(
    students: list[dict[str, Any]],
    score_long: list[dict[str, Any]],
    class_averages: list[dict[str, Any]],
    target_class: str,
) -> list[dict[str, Any]]:
    """学籍1科目均分估算（单学科化，阶段7）：只按科目做均分估算排名，
    不再做主三门/五门总分均分估算（TotalScore 已退役）。
    """
    rows: list[dict[str, Any]] = []
    actual_class_rows = [row for row in class_averages if row.get("is_actual_class")]
    exams = available_exams(students)
    for exam in exams:
        target_student_ids = {
            row["student_id"]
            for row in students
            if row["exam"] == exam and row["class"] == target_class and row["xueji"] == "1"
        }
        if not target_student_ids:
            continue
        exam_class_rows = [row for row in actual_class_rows if row["exam"] == exam]
        target_official = next((row for row in exam_class_rows if row["class"] == target_class), None)
        for subject in SUBJECT_COLS:
            values = [
                row["score"]
                for row in score_long
                if row["exam"] == exam and row["student_id"] in target_student_ids and row["subject"] == subject and row.get("score") is not None
            ]
            if not values:
                continue
            avg = mean(values)
            parallel_comparators = [
                row.get(f"{subject}_avg")
                for row in exam_class_rows
                if row.get("class_type") == "平行班" and row.get("class") != target_class
            ]
            grade_comparators = [
                row.get(f"{subject}_avg")
                for row in exam_class_rows
                if row.get("class") != target_class
            ]
            rows.append({
                "exam": exam,
                "exam_order": EXAM_ORDER.get(exam, 999),
                "metric": subject,
                "metric_type": "科目",
                "xueji1_count": len(target_student_ids),
                "xueji1_avg": round(avg, 2),
                "official_class_avg": target_official.get(f"{subject}_avg") if target_official else None,
                "official_rank": "",
                "estimated_parallel_rank": estimated_rank(avg, parallel_comparators),
                "estimated_grade_rank": estimated_rank(avg, grade_comparators),
                "note": "本班学籍1均分与其他班现有均分比较，属于估算。",
            })
    rows.sort(key=lambda row: (row["exam_order"], row["metric_type"], row["metric"]))
    return rows


def build_class_overview(class_averages: list[dict[str, Any]], target_class: str) -> list[dict[str, Any]]:
    rows = [row for row in class_averages if row.get("is_actual_class") and row.get("class") == target_class]
    rows.sort(key=lambda row: row.get("exam_order", 999))
    return rows


def build_history_rows(class_overview: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """本班历史科目均分趋势（单学科化，阶段7）：只输出科目均分行，
    不再输出主三门/五门总分均分行。
    """
    rows: list[dict[str, Any]] = []
    for row in class_overview:
        for subject in SUBJECT_COLS:
            value = row.get(f"{subject}_avg")
            if value not in (None, 0):
                rows.append({
                    "exam": row["exam"],
                    "exam_order": row["exam_order"],
                    "metric_type": "科目均分",
                    "metric": subject,
                    "value": value,
                    "rank": "",
                })
    return rows


def build_analysis(parsed: dict[str, Any], target_class: str) -> dict[str, Any]:
    """构建分析结果（单学科化，阶段7）。

    不再有 student_trends / focus_students / subject_communication /
    target_rank_bands / uploaded_rank_band_summary（均依赖总分名次）。
    只保留班级总览、平行班对比、学籍1科目均分估算、本班历史科目均分趋势。
    """
    students = parsed["students"]
    score_long = parsed["score_long"]
    class_averages = parsed["class_averages"]
    class_overview = build_class_overview(class_averages, target_class)
    xueji1_analysis = build_xueji1_analysis(students, score_long, class_averages, target_class)
    actual_class_averages = [row for row in class_averages if row.get("is_actual_class")]
    parallel = [row for row in actual_class_averages if row.get("class_type") == "平行班"]
    exams = available_exams(students + class_averages)
    latest = latest_exam(students + class_averages)
    previous = previous_exam_name(exams, latest)
    warnings = list(parsed["warnings"])
    if not students:
        warnings.append("没有找到学生成绩表，单科成绩分析不可用。")
    if not class_averages:
        warnings.append("没有找到班级均分表，班级对比不可用。")
    if class_overview and latest:
        latest_target = next((row for row in class_overview if row["exam"] == latest), None)
        if latest_target is None:
            warnings.append(f"最新考试 {latest} 没有目标班 {target_class} 的班级均分行。")
    return {
        "target_class": target_class,
        "exams": exams,
        "latest_exam": latest,
        "previous_exam": previous,
        "warnings": warnings,
        "students": students,
        "score_long": score_long,
        "class_averages": class_averages,
        "class_overview": class_overview,
        "parallel_comparison": parallel,
        "grade_reference": actual_class_averages,
        "history_rows": build_history_rows(class_overview),
        "xueji1_analysis": xueji1_analysis,
    }


def json_safe(value: Any) -> Any:
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                fieldnames.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: json_safe(row.get(key, "")) for key in fieldnames})


def display_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 4)
    return value


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]] | list[list[Any]]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["说明"])
        ws.append(["无数据"])
        return
    if isinstance(rows[0], dict):
        fieldnames: list[str] = []
        seen: set[str] = set()
        for row in rows:  # type: ignore[assignment]
            for key in row.keys():  # type: ignore[union-attr]
                if key not in seen:
                    seen.add(key)
                    fieldnames.append(key)
        ws.append(fieldnames)
        for row in rows:  # type: ignore[assignment]
            ws.append([display_value(row.get(key)) for key in fieldnames])  # type: ignore[union-attr]
    else:
        for row in rows:  # type: ignore[assignment]
            ws.append([display_value(value) for value in row])
    style_sheet(ws)


def style_sheet(ws: Any) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    if ws.max_row > 1 and ws.max_column > 1:
        ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_len = 8
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            max_len = max(max_len, min(len(clean(cell.value)), 50))
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = max_len + 2


def make_notes_rows(analysis: dict[str, Any]) -> list[list[Any]]:
    return [
        ["项目", "说明"],
        ["目标班级", analysis["target_class"]],
        ["考试顺序", " -> ".join(analysis["exams"])],
        ["单科趋势", "按年级百分位判断，未考科目不参与比较；百分位降低为进步。"],
        ["学籍1估算", "本班学籍1科目均分与其他班现有均分比较，属于估算，不是全年级精确重排。"],
        ["数据包", "CSV/JSON保留全量明细，支持后续聊天追问。"],
    ]


def write_workbook(path: Path, analysis: dict[str, Any]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    add_sheet(wb, "口径说明", make_notes_rows(analysis))
    add_sheet(wb, "数据质量检查", [{"warning": warning} for warning in analysis["warnings"]])
    add_sheet(wb, "班级总览", analysis["class_overview"])
    add_sheet(wb, "平行班对比", analysis["parallel_comparison"])
    add_sheet(wb, "全年级参考", analysis["grade_reference"])
    add_sheet(wb, "本班历史趋势", analysis["history_rows"])
    add_sheet(wb, "学籍1分析", analysis["xueji1_analysis"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def summarize_class_change(class_overview: list[dict[str, Any]], latest: str | None, previous: str | None) -> list[str]:
    """班级科目均分变化摘要（单学科化，阶段7）：只输出科目均分变化，
    不再输出主三门/五门总分均分变化。
    """
    if not latest:
        return ["未识别到最新考试。"]
    latest_row = next((row for row in class_overview if row["exam"] == latest), None)
    previous_row = next((row for row in class_overview if row["exam"] == previous), None) if previous else None
    if not latest_row:
        return [f"最新考试 {latest} 未找到目标班级均分。"]
    lines: list[str] = []
    for subject in SUBJECT_COLS:
        avg = latest_row.get(f"{subject}_avg")
        if avg in (None, 0):
            continue
        line = f"- {latest} {subject}均分 {avg}"
        if previous_row and previous_row.get(f"{subject}_avg") not in (None, 0):
            avg_delta = float(avg) - float(previous_row.get(f"{subject}_avg"))
            line += f"，较{previous}变化 {avg_delta:.2f}"
        lines.append(line)
    return lines or [f"{latest} 暂无可用科目均分。"]


def summarize_xueji1(rows: list[dict[str, Any]], latest: str | None) -> list[str]:
    """学籍1科目均分估算摘要（单学科化，阶段7）：只输出科目行，
    不再输出主三门/五门总分行。
    """
    if not latest:
        return ["无最新考试。"]
    latest_rows = [
        row
        for row in rows
        if row["exam"] == latest
        and row["metric_type"] == "科目"
    ]
    if not latest_rows:
        return ["最新考试暂无学籍1科目均分估算。"]
    lines = []
    for row in latest_rows:
        lines.append(
            f"- {row['metric']} 学籍1均分 {row['xueji1_avg']}，平行班估算排名 {row['estimated_parallel_rank']}，全年级估算排名 {row['estimated_grade_rank']}（估算）"
        )
    return lines


def write_summary(path: Path, analysis: dict[str, Any]) -> None:
    latest = analysis["latest_exam"]
    previous = analysis["previous_exam"]
    lines = [
        "# 成绩分析摘要",
        "",
        f"- 目标班级: {analysis['target_class']}班",
        f"- 考试范围: {' -> '.join(analysis['exams'])}",
        f"- 最新考试: {latest or '未识别'}",
        "",
        "## 数据质量",
    ]
    if analysis["warnings"]:
        lines.extend([f"- {warning}" for warning in analysis["warnings"]])
    else:
        lines.append("- 未发现关键数据质量问题。")
    lines.extend(["", "## 班级科目均分"])
    lines.extend(summarize_class_change(analysis["class_overview"], latest, previous))
    lines.extend(["", "## 学籍1科目估算"])
    lines.extend(summarize_xueji1(analysis["xueji1_analysis"], latest))
    lines.extend([
        "",
        "## 口径提醒",
        "- 单科趋势按年级百分位判断，未考科目不参与比较。",
        "- 学籍1排名变化是本班学籍1均分与其他班现有均分比较的估算，不是精确重排。",
    ])
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_data_package(out_dir: Path, analysis: dict[str, Any]) -> dict[str, str]:
    data_dir = out_dir / "data"
    outputs = {
        "score_long": data_dir / "score_long.csv",
        "class_averages": data_dir / "class_averages.csv",
    }
    write_csv(outputs["score_long"], analysis["score_long"])
    write_csv(outputs["class_averages"], analysis["class_averages"])
    json_path = data_dir / "analysis.json"
    serializable = {
        key: value
        for key, value in analysis.items()
        if key
        in {
            "target_class",
            "exams",
            "latest_exam",
            "previous_exam",
            "warnings",
            "class_overview",
            "xueji1_analysis",
        }
    }
    json_path.write_text(json.dumps(serializable, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    result = {name: str(path) for name, path in outputs.items()}
    result["analysis_json"] = str(json_path)
    return result


def run_analysis(paths: list[Path], out_dir: Path, target_class: str | None) -> dict[str, Any]:
    parsed = parse_inputs(paths)
    resolved_target = target_class or most_common_target_class(parsed["students"])
    analysis = build_analysis(parsed, resolved_target)
    out_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = out_dir / "成绩分析.xlsx"
    summary_path = out_dir / "成绩分析摘要.md"
    write_workbook(workbook_path, analysis)
    write_summary(summary_path, analysis)
    data_outputs = write_data_package(out_dir, analysis)
    html_outputs: dict[str, str] = {}
    if build_learning_profile_outputs is not None:
        try:
            html_outputs = build_learning_profile_outputs(
                data_dir=out_dir / "data",
                out_dir=out_dir,
                target_class=resolved_target,
            )
        except Exception as exc:
            analysis["warnings"].append(f"学生学情HTML生成失败: {exc}")
    return {
        "workbook": str(workbook_path),
        "summary": str(summary_path),
        "data": data_outputs,
        "student_learning_profile": html_outputs,
        "warnings": analysis["warnings"],
        "target_class": resolved_target,
        "exams": analysis["exams"],
    }


def existing_paths(items: list[str]) -> list[Path]:
    paths = [Path(item).expanduser() for item in items]
    missing = [str(path) for path in paths if not path.exists()]
    if missing:
        raise SystemExit("Missing input files:\n" + "\n".join(missing))
    return paths


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect and analyze exam score workbooks.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    inspect_parser = subparsers.add_parser("inspect", help="Inspect workbook structures.")
    inspect_parser.add_argument("--inputs", nargs="+", required=True, help="Input Excel workbooks.")
    inspect_parser.add_argument("--json-out", help="Optional JSON output path.")

    run_parser = subparsers.add_parser("run", help="Generate workbook, summary, and data package.")
    run_parser.add_argument("--inputs", nargs="+", required=True, help="Input Excel workbooks.")
    run_parser.add_argument("--out", required=True, help="Output directory.")
    run_parser.add_argument("--target-class", default=None, help="Target class, default auto-detects from student scores.")
    run_parser.add_argument("--json-out", help="Optional JSON output path for run manifest.")

    args = parser.parse_args()

    if args.command == "inspect":
        result = inspect_files(existing_paths(args.inputs))
        text = json.dumps(result, ensure_ascii=False, indent=2, default=str)
        if args.json_out:
            Path(args.json_out).expanduser().write_text(text + "\n", encoding="utf-8")
        print(text)
    elif args.command == "run":
        result = run_analysis(existing_paths(args.inputs), Path(args.out).expanduser(), args.target_class)
        text = json.dumps(result, ensure_ascii=False, indent=2, default=str)
        if args.json_out:
            Path(args.json_out).expanduser().write_text(text + "\n", encoding="utf-8")
        print(text)


if __name__ == "__main__":
    main()
