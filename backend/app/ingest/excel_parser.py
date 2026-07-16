from __future__ import annotations

import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# 解析脚本已并入本包（原先依赖上级目录 exam-score-analysis/scripts，迁移后会断），
# 改为包内导入，使 webapp 可独立部署。
from app.ingest.analyze_exam_scores import (
    classify_workbook,
    parse_class_averages,
    parse_student_scores,
)


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip().replace("\n", " ")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = clean(value).replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return None


def percentile(value: Any) -> float | None:
    value_num = number(value)
    if value_num is None:
        return None
    return value_num / 100 if value_num > 1 else value_num


def integer(value: Any) -> int | None:
    value_num = number(value)
    if value_num is None:
        return None
    return int(round(value_num))


def class_num(value: Any) -> int | None:
    text = clean(value)
    if not text.isdigit():
        return None
    return int(text)


# 教学班/走班列的可选表头名（关键字匹配，不写死列号，降风险）
TEACHING_CLASS_HEADERS = {
    "教学班", "走班", "走班班级", "教学班级", "选科班", "选科班级", "走班教学班", "行政教学班"
}


def parse_class_label(value: Any) -> str | None:
    """解析教学班标签：去空白；数字也保留为字符串（"5"）；空→None。
    区别于只收数字的 class_num()：本函数对 '物A1'/'5' 一视同仁。"""
    text = clean(value)
    if not text:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if float(value).is_integer():
            return str(int(value))
    return text


def detect_class_label_col(headers: list[str]) -> int | None:
    """从一组表头（1-based 对应 col=index+1）里探测教学班列，返回 1-based 列号或 None。"""
    for idx, header in enumerate(headers):
        if header and header.replace(" ", "") in TEACHING_CLASS_HEADERS:
            return idx + 1
    return None


def fill_merged_headers(values: list[str]) -> list[str]:
    filled = []
    current = ""
    for value in values:
        if value:
            current = value
        filled.append(current)
    return filled


def normalize_grade1_student_scores(file_path: Path, info: dict[str, Any]) -> dict:
    students, subject_scores = parse_student_scores(file_path, info)
    return {
        "kind": "student_scores",
        "students": [
            {
                **student,
                "class_num": class_num(student.get("class")),
                "xueji": integer(student.get("xueji")),
            }
            for student in students
        ],
        "subject_scores": [
            {
                "student_id": row["student_id"],
                "class_num": class_num(row.get("class")),
                "xueji": integer(row.get("xueji")),
                "name": row.get("name"),
                "subject": row["subject"],
                "raw_score": row.get("score"),
                "grade_score": None,
                "grade_percentile": row.get("grade_percentile"),
            }
            for row in subject_scores
        ],
    }


GRADE1_SUBJECTS = ["语文", "数学", "英语", "物理", "化学", "生物", "政治", "历史", "地理"]


def is_grade1_student_sheet(path: Path) -> bool:
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    headers = [clean(ws.cell(2, col).value) for col in range(1, 5)]
    return headers == ["学号", "班级", "学籍", "姓名"] and clean(ws.cell(3, 6).value) == "年级百分位"


def parse_excel_grade1_student_scores(path: Path) -> dict:
    """解析高一学生成绩明细表，包含年级百分位、学籍排名、年级排名。"""
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]

    top_headers = fill_merged_headers([clean(ws.cell(2, col).value) for col in range(1, ws.max_column + 1)])
    sub_headers = [clean(ws.cell(3, col).value) for col in range(1, ws.max_column + 1)]
    class_label_col = detect_class_label_col(top_headers) or detect_class_label_col(sub_headers)
    subject_cols: dict[str, dict[str, int]] = {subject: {} for subject in GRADE1_SUBJECTS}

    for col, (top, sub) in enumerate(zip(top_headers, sub_headers), start=1):
        normalized_top = top.replace(" ", "")
        normalized_sub = sub.replace(" ", "")
        if normalized_top in subject_cols:
            if normalized_sub == "分数":
                subject_cols[normalized_top]["score"] = col
            elif normalized_sub == "年级百分位":
                subject_cols[normalized_top]["percentile"] = col

    students: list[dict[str, Any]] = []
    subject_scores: list[dict[str, Any]] = []

    for row_idx in range(4, ws.max_row + 1):
        student_id = clean(ws.cell(row_idx, 1).value)
        if not student_id:
            continue

        row_class_num = class_num(ws.cell(row_idx, 2).value)
        row_class_label = parse_class_label(ws.cell(row_idx, class_label_col).value) if class_label_col else None
        row_xueji = integer(ws.cell(row_idx, 3).value)
        name = clean(ws.cell(row_idx, 4).value)
        students.append(
            {
                "student_id": student_id,
                "class_num": row_class_num,
                "class_label": row_class_label,
                "xueji": row_xueji,
                "name": name,
                "grade": 1,
                "source_file": path.name,
            }
        )

        for subject, cols in subject_cols.items():
            raw_score = number(ws.cell(row_idx, cols["score"]).value) if cols.get("score") else None
            grade_percentile = percentile(ws.cell(row_idx, cols["percentile"]).value) if cols.get("percentile") else None
            if raw_score is None and grade_percentile is None:
                continue
            subject_scores.append(
                {
                    "student_id": student_id,
                    "class_num": row_class_num,
                    "class_label": row_class_label,
                    "xueji": row_xueji,
                    "name": name,
                    "subject": subject,
                    "raw_score": raw_score,
                    "grade_score": None,
                    "grade_percentile": grade_percentile,
                }
            )

    return {
        "kind": "student_scores",
        "students": students,
        "subject_scores": subject_scores,
    }


def normalize_grade1_class_averages(file_path: Path, info: dict[str, Any]) -> dict:
    rows = parse_class_averages(file_path, info)
    class_averages = []
    subjects = ["语文", "数学", "英语", "物理", "化学", "生物", "政治", "历史", "地理"]
    for row in rows:
        parsed_class_num = class_num(row.get("class"))
        if parsed_class_num is None:
            continue
        class_averages.append(
            {
                "class_type": row.get("class_type"),
                "class_num": parsed_class_num,
                "teacher_name": row.get("teacher"),
                "subject_averages": {
                    subject: row.get(f"{subject}_avg")
                    for subject in subjects
                    if row.get(f"{subject}_avg") is not None
                },
            }
        )
    return {"kind": "class_averages", "class_averages": class_averages}


def parse_excel_grade1(file_path: str) -> dict:
    """解析高一 Excel,复用旧 Skill parser 并正规化成 WebApp 入库字段 - """
    path = Path(file_path)
    if is_grade1_student_sheet(path):
        return parse_excel_grade1_student_scores(path)

    info = classify_workbook(path)

    if info["type"] == "student_scores":
        return normalize_grade1_student_scores(path, info)
    if info["type"] == "class_averages":
        return normalize_grade1_class_averages(path, info)
    return {"kind": "unknown", "workbook_type": info["type"]}


GRADE23_SUBJECTS = [
    ("语文", 5, None, 27),
    ("数学", 6, None, 28),
    ("英语", 7, None, 29),
    ("物理", 8, 9, None),
    ("化学", 10, 11, None),
    ("生物", 12, 13, None),
    ("政治", 14, 15, None),
    ("历史", 16, 17, None),
    ("地理", 18, 19, None),
]

GRADE23_CLASS_AVG_BASE_SUBJECTS = {"语文", "数学", "英语"}
GRADE23_CLASS_AVG_ELECTIVE_SUBJECTS = {"物理", "化学", "生物", "政治", "历史", "地理", "物", "化", "生", "政", "史", "地"}
GRADE23_SHORT_SUBJECTS = {
    "物": "物理",
    "化": "化学",
    "生": "生物",
    "政": "政治",
    "史": "历史",
    "地": "地理",
}


def is_grade23_student_sheet(path: Path) -> bool:
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    headers = [clean(ws.cell(2, col).value) for col in range(1, 5)]
    return headers == ["学号", "班级", "学籍", "姓名"] and clean(ws.cell(2, 20).value) == "+3总分"


def find_grade23_class_average_header(ws: Any) -> int | None:
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        values = [clean(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, 8) + 1)]
        normalized = [value.replace(" ", "") for value in values]
        if any(value in {"班型", "班级类型"} for value in normalized) and "班级" in normalized and "班主任" in normalized:
            return row_idx
    return None


def parse_excel_grade23_class_averages(path: Path) -> dict:
    """解析高二/高三 3+3 班级均分表 Excel - """
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    header_row = find_grade23_class_average_header(ws)
    if header_row is None:
        return {"kind": "unknown", "message": "未识别为高二/高三班级均分表"}

    top_headers = fill_merged_headers([clean(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)])
    sub_headers = [clean(ws.cell(header_row + 1, col).value) for col in range(1, ws.max_column + 1)]
    class_type_col = None
    class_col = None
    teacher_col = None
    subject_cols: dict[str, int] = {}

    for col, (top, sub) in enumerate(zip(top_headers, sub_headers), start=1):
        normalized_top = top.replace(" ", "")
        normalized_sub = sub.replace(" ", "")
        normalized_sub_type = normalized_sub.removesuffix("分")

        if normalized_top in {"班型", "班级类型"}:
            class_type_col = col
        elif normalized_top == "班级":
            class_col = col
        elif normalized_top == "班主任":
            teacher_col = col
        elif normalized_top in GRADE23_CLASS_AVG_BASE_SUBJECTS:
            subject_cols[normalized_top] = col
        elif normalized_top.endswith("原始") or normalized_top.endswith("等级") or normalized_top.endswith("原始分") or normalized_top.endswith("等级分"):
            if normalized_top.endswith("原始分"):
                suffix = "原始"
                subject = normalized_top[:-3]
            elif normalized_top.endswith("等级分"):
                suffix = "等级"
                subject = normalized_top[:-3]
            elif normalized_top.endswith("原始"):
                suffix = "原始"
                subject = normalized_top[:-2]
            else:
                suffix = "等级"
                subject = normalized_top[:-2]
            if subject in GRADE23_CLASS_AVG_ELECTIVE_SUBJECTS:
                subject_cols[f"{GRADE23_SHORT_SUBJECTS.get(subject, subject)}_{suffix}"] = col
        elif normalized_top in GRADE23_CLASS_AVG_ELECTIVE_SUBJECTS and normalized_sub_type in {"原始", "等级"}:
            subject = GRADE23_SHORT_SUBJECTS.get(normalized_top, normalized_top)
            subject_cols[f"{subject}_{normalized_sub_type}"] = col

    if class_col is None:
        return {"kind": "unknown", "message": "班级均分表缺少班级列"}

    class_label_col = detect_class_label_col(top_headers) or detect_class_label_col(sub_headers)

    class_averages = []
    current_class_type = ""
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_class_num = class_num(ws.cell(row_idx, class_col).value)
        if row_class_num is None:
            class_type_value = clean(ws.cell(row_idx, class_type_col).value) if class_type_col else ""
            if class_type_value:
                current_class_type = class_type_value
            continue

        class_type_value = clean(ws.cell(row_idx, class_type_col).value) if class_type_col else ""
        if class_type_value:
            current_class_type = class_type_value

        row_class_label = parse_class_label(ws.cell(row_idx, class_label_col).value) if class_label_col else None

        subject_averages = {
            subject: value
            for subject, col in subject_cols.items()
            if (value := number(ws.cell(row_idx, col).value)) is not None
        }

        class_averages.append(
            {
                "class_type": current_class_type or None,
                "class_num": row_class_num,
                "class_label": row_class_label,
                "teacher_name": clean(ws.cell(row_idx, teacher_col).value) if teacher_col else None,
                "subject_averages": subject_averages,
            }
        )

    return {"kind": "class_averages", "class_averages": class_averages}


def parse_excel_grade23(file_path: str, grade: int) -> dict:
    """解析高二/高三 3+3 Excel:学生成绩明细表或班级均分表 - """
    path = Path(file_path)
    if not is_grade23_student_sheet(path):
        return parse_excel_grade23_class_averages(path)

    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    grade23_headers = [clean(ws.cell(2, col).value) for col in range(1, ws.max_column + 1)]
    class_label_col = detect_class_label_col(grade23_headers)
    students: list[dict[str, Any]] = []
    subject_scores: list[dict[str, Any]] = []

    for row_idx in range(4, ws.max_row + 1):
        student_id = clean(ws.cell(row_idx, 1).value)
        if not student_id:
            continue

        row_class_num = class_num(ws.cell(row_idx, 2).value)
        row_class_label = parse_class_label(ws.cell(row_idx, class_label_col).value) if class_label_col else None
        row_xueji = integer(ws.cell(row_idx, 3).value)
        name = clean(ws.cell(row_idx, 4).value)
        students.append(
            {
                "student_id": student_id,
                "class_num": row_class_num,
                "class_label": row_class_label,
                "xueji": row_xueji,
                "name": name,
                "grade": grade,
                "source_file": path.name,
            }
        )

        for subject, raw_col, grade_col, pct_col in GRADE23_SUBJECTS:
            raw_score = number(ws.cell(row_idx, raw_col).value)
            grade_score = number(ws.cell(row_idx, grade_col).value) if grade_col else None
            grade_percentile = percentile(ws.cell(row_idx, pct_col).value) if pct_col else None
            if raw_score is None and grade_score is None and grade_percentile is None:
                continue
            subject_scores.append(
                {
                    "student_id": student_id,
                    "class_num": row_class_num,
                    "class_label": row_class_label,
                    "xueji": row_xueji,
                    "name": name,
                    "subject": subject,
                    "raw_score": raw_score,
                    "grade_score": grade_score,
                    "grade_percentile": grade_percentile,
                }
            )

    return {
        "kind": "student_scores",
        "students": students,
        "subject_scores": subject_scores,
    }
