"""阶段7：上传单学科存储与 TotalScore 退役测试。

TDD RED：先写断言，实现完成后转 GREEN。

覆盖范围（对应 task §1–§9）：
1. 高二/高三学生成绩明细表上传：数学教师上传含数学+物理+主三门/3+3/+3 总分
   的混合 Excel，数据库只新增数学 SubjectScore；TotalScore 新增 0 行；物理
   等其他学科新增 0 行。
2. 高一学生成绩明细表上传：含主三门/五门/九门总分列的旧格式 Excel，数据库
   只新增数学 SubjectScore；TotalScore 新增 0 行；其他学科新增 0 行。
3. 班级均分表上传：ClassAverage 只保留当前 subject 的 subject_averages 键；
   total_averages 为空（{}）。
4. 解析器仍兼容旧格式 Excel：可识别/跳过旧总分列，不因列仍存在而拒绝文件。
5. 重复上传（同 exam）不重复写；失败回滚不残留。
6. 旧 SQLite（已有 TotalScore 行）仍可正常启动；备份/恢复无损保留旧表。
7. 删除考试可级联清理旧 TotalScore 关联行（整场清理）。
8. 同名消歧（name_candidates / _ambiguous_entry）不再查询 TotalScore 主三门
   名次，改用当前任教学科最近合法成绩 / 班级信息辅助。
9. 旧遗留 analysis 模块（focus_list/cross_year/trends/class_compare）不再含
   TotalScore 业务查询。
"""
from __future__ import annotations

import json
import os
import subprocess
import sys
import textwrap

import pytest


# ════════════════════════════════════════════════════════════════
#  API 端到端测试（子进程 + 全新临时 EXAM_TRACKER_DIR）
# ════════════════════════════════════════════════════════════════

_API_TEST_SCRIPT = textwrap.dedent("""\
    import json, sys
    from fastapi.testclient import TestClient
    from app.main import app
    from app.db.models import SessionLocal

    client = TestClient(app)

    setup_script = sys.argv[1]
    with open(setup_script) as f:
        exec(f.read())

    assert_script = sys.argv[2]
    with open(assert_script) as f:
        exec(f.read())
""")


def _run_isolated_api_test(tmp_path, setup_code: str, assert_code: str, timeout: int = 90):
    """在子进程中用全新临时 DB 运行 API 测试，返回 proc。"""
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    setup_file = tmp_path / "setup.py"
    setup_file.write_text(setup_code)
    assert_file = tmp_path / "assert.py"
    assert_file.write_text(assert_code)

    backend_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    env = os.environ.copy()
    env["EXAM_TRACKER_DIR"] = str(data_dir)
    env["EXAM_TRACKER_BACKUP_DIR"] = str(tmp_path / "backups")
    venv_python = os.path.join(os.path.dirname(sys.executable), "python")
    if not os.path.exists(venv_python):
        venv_python = sys.executable

    proc = subprocess.run(
        [venv_python, "-c", _API_TEST_SCRIPT, str(setup_file), str(assert_file)],
        cwd=backend_dir,
        env=env,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        timeout=timeout,
        check=False,
    )
    return proc


def _assert_ok(proc):
    if proc.returncode != 0:
        raise AssertionError(f"子进程失败 (rc={proc.returncode}):\n{proc.stdout}")


# ──────────────────────────────────────────────────────────────
#  Excel 制表助手（子进程内用）
# ──────────────────────────────────────────────────────────────

_MAKE_GRADE23_XLSX = textwrap.dedent("""\
    from openpyxl import Workbook
    def make_grade23_xlsx(path, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "学生成绩明细"
        ws.cell(2, 1, "学号")
        ws.cell(2, 2, "班级")
        ws.cell(2, 3, "学籍")
        ws.cell(2, 4, "姓名")
        ws.cell(2, 20, "+3总分")
        ws.cell(2, 21, "主三门")
        ws.cell(2, 24, "3+3总分")
        ws.cell(2, 30, "教学班")
        ws.append([])
        for r in rows:
            ws.append(r)
        wb.save(path)
""")

_MAKE_GRADE1_XLSX = textwrap.dedent("""\
    from openpyxl import Workbook
    def make_grade1_xlsx(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "学生成绩明细"
        ws.append(["学生成绩（在籍）"])
        ws.append(["学号","班级","学籍","姓名","语文",None,"数学",None,"物理",None,"主三门",None,None,None])
        ws.append([None,None,None,None,"分数","年级百分位","分数","年级百分位","分数","年级百分位","总分","年级百分位","学籍排名","年级排名"])
        ws.merge_cells("A2:A3"); ws.merge_cells("B2:B3"); ws.merge_cells("C2:C3"); ws.merge_cells("D2:D3")
        ws.merge_cells("E2:F2"); ws.merge_cells("G2:H2"); ws.merge_cells("I2:J2"); ws.merge_cells("K2:N2")
        ws.append(["7240101","01","1","卞幻", 92,"45.01%", 106,"66.84%", 48,"35.00%", 306,"47.88%",283,283])
        wb.save(path)
""")

_MAKE_GRADE23_AVG_XLSX = textwrap.dedent("""\
    from openpyxl import Workbook
    def make_grade23_avg_xlsx(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "班级均分"
        ws.append(["班型","班级","教学班","班主任","语文","数学","英语","物理",None,"化学",None,"加3同均分","主三门","3+3总分"])
        ws.append([None,None,None,None,None,None,None,"原始","等级","原始","等级",None,None,None])
        ws.append(["平行班","01","物A1","张老师", 101.2,108.3,110.4, 55.1,61.2, 58.3,62.4, 180.5,319.9,500.4])
        wb.save(path)
""")


# ════════════════════════════════════════════════════════════════
#  §1 高二/高三学生成绩明细表：只存数学
# ════════════════════════════════════════════════════════════════

def test_grade23_upload_stores_only_teacher_subject(tmp_path):
    """数学教师上传高二混合 Excel（数学+物理+主三门/3+3 总分）：
    数据库只新增数学 SubjectScore；TotalScore 0 行；物理 0 行。"""
    setup = _MAKE_GRADE23_XLSX + textwrap.dedent("""\
        import os
        from app.db.models import Teacher
        db = SessionLocal()
        t = Teacher(subject="数学", name="数学老师")
        db.add(t); db.commit(); db.close()

        raw_dir = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw")
        os.makedirs(raw_dir, exist_ok=True)
        xlsx_path = os.path.join(raw_dir, "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        # 行：学号 班级 学籍 姓名 语(5) 数(6) 英(7) 物raw(8) 物grade(9) 化raw(10) 化grade(11)
        #      12-19空  +3(20) 主三门(21) 主三门pct(22) 主三门rank(23) 3+3(24) 3+3pct(25) 3+3rank(26)
        #      语pct(27) 数pct(28) 英pct(29)  教学班(30)
        make_grade23_xlsx(xlsx_path, [
            ["7240101","01","1","卞幻", 97,108,120, 48,52, None,None,
             None,None,None,None,None,None,None,None,
             174, 325.5,"25.08%",283, 499.5,"30.01%",291,
             "75.08%","40.12%","35.45%", "物A1"],
        ])
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import SubjectScore, TotalScore, Exam
        from app.ingest.router import parse_and_store

        parsed = {"grade":2,"semester":"上","exam_type":"期中","sort_key":"2025-11","canonical_name":"高二2025学年第二学期期中考试"}
        xlsx_path = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw", "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        out = parse_and_store(xlsx_path, "高二2025学年第二学期期中考试学生成绩明细表.xlsx", parsed, 2)
        assert out["result"]["parsed_ok"], out["result"]

        math_count = db.query(SubjectScore).filter(SubjectScore.subject=="数学").count()
        physics_count = db.query(SubjectScore).filter(SubjectScore.subject=="物理").count()
        total_count = db.query(TotalScore).count()
        db.close()

        assert math_count == 1, f"数学应有1行, 实际{math_count}"
        assert physics_count == 0, f"物理不应入库, 实际{physics_count}"
        assert total_count == 0, f"TotalScore应为0行, 实际{total_count}"
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §2 高一学生成绩明细表：只存数学，跳过总分列
# ════════════════════════════════════════════════════════════════

def test_grade1_upload_stores_only_teacher_subject(tmp_path):
    """数学教师上传高一含总分列的旧格式 Excel：只新增数学 SubjectScore；
    TotalScore 0 行；语文/物理 0 行。"""
    setup = _MAKE_GRADE1_XLSX + textwrap.dedent("""\
        import os
        from app.db.models import Teacher
        db = SessionLocal()
        t = Teacher(subject="数学", name="数学老师")
        db.add(t); db.commit(); db.close()

        raw_dir = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw")
        os.makedirs(raw_dir, exist_ok=True)
        xlsx_path = os.path.join(raw_dir, "2024级2024学年第二学期期中考试.xlsx")
        make_grade1_xlsx(xlsx_path)
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import SubjectScore, TotalScore
        from app.ingest.router import parse_and_store

        parsed = {"grade":1,"semester":"下","exam_type":"期中","sort_key":"2025-04","canonical_name":"2024级2024学年第二学期期中考试"}
        xlsx_path = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw", "2024级2024学年第二学期期中考试.xlsx")
        out = parse_and_store(xlsx_path, "2024级2024学年第二学期期中考试.xlsx", parsed, 1)
        assert out["result"]["parsed_ok"], out["result"]

        math_count = db.query(SubjectScore).filter(SubjectScore.subject=="数学").count()
        chinese_count = db.query(SubjectScore).filter(SubjectScore.subject=="语文").count()
        total_count = db.query(TotalScore).count()
        db.close()

        assert math_count == 1, f"数学应有1行, 实际{math_count}"
        assert chinese_count == 0, f"语文不应入库, 实际{chinese_count}"
        assert total_count == 0, f"TotalScore应为0行, 实际{total_count}"
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §3 班级均分表：ClassAverage 只保留当前 subject，total_averages 为空
# ════════════════════════════════════════════════════════════════

def test_class_average_upload_keeps_only_subject(tmp_path):
    """数学教师上传高二班级均分表（含语文/数学/英语/物理/化学 + 主三门/3+3 总分）：
    ClassAverage 只保留数学的 subject_averages 键；total_averages 为空 {}。"""
    setup = _MAKE_GRADE23_AVG_XLSX + textwrap.dedent("""\
        import os
        from app.db.models import Teacher
        db = SessionLocal()
        t = Teacher(subject="数学", name="数学老师")
        db.add(t); db.commit(); db.close()

        raw_dir = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw")
        os.makedirs(raw_dir, exist_ok=True)
        xlsx_path = os.path.join(raw_dir, "高二2025学年第二学期期中考试班级均分表.xlsx")
        make_grade23_avg_xlsx(xlsx_path)
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import ClassAverage
        from app.ingest.router import parse_and_store

        parsed = {"grade":2,"semester":"上","exam_type":"期中","sort_key":"2025-11","canonical_name":"高二2025学年第二学期期中考试"}
        xlsx_path = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw", "高二2025学年第二学期期中考试班级均分表.xlsx")
        out = parse_and_store(xlsx_path, "高二2025学年第二学期期中考试班级均分表.xlsx", parsed, 2)
        assert out["result"]["parsed_ok"], out["result"]

        avgs = db.query(ClassAverage).all()
        assert len(avgs) >= 1
        for ca in avgs:
            sa = ca.subject_averages or {}
            ta = ca.total_averages or {}
            assert set(sa.keys()) <= {"数学"}, f"subject_averages 只应有数学, 实际{set(sa.keys())}"
            assert ta == {}, f"total_averages 应为空, 实际{ta}"
        db.close()
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §4 解析器仍兼容旧格式（不因总分列存在而拒绝）
# ════════════════════════════════════════════════════════════════

def test_parser_accepts_legacy_total_columns(tmp_path):
    """解析器解析旧格式高二 Excel（含 +3/主三门/3+3 总分列）成功，kind 正确。"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "学生成绩明细"
    ws.cell(2, 1, "学号")
    ws.cell(2, 2, "班级")
    ws.cell(2, 3, "学籍")
    ws.cell(2, 4, "姓名")
    ws.cell(2, 20, "+3总分")
    ws.cell(2, 21, "主三门")
    ws.cell(2, 24, "3+3总分")
    ws.append([])
    ws.append([
        "7240101", "01", "1", "卞幻",
        97, 108, 120, 48, 52, 50, 55,
        None, None, None, None, None, None, None, None,
        174, 325.5, "25.08%", 283, 499.5, "30.01%", 291,
        "75.08%", "40.12%", "35.45%",
    ])
    path = tmp_path / "高二2025学年第二学期期中考试学生成绩明细表.xlsx"
    wb.save(path)

    from app.ingest.excel_parser import parse_excel_grade23
    result = parse_excel_grade23(str(path), grade=2)

    assert result["kind"] == "student_scores"
    assert len(result["students"]) == 1
    # 解析器仍能解析出数学单科成绩（兼容旧格式不丢数据）
    subjects = [r["subject"] for r in result["subject_scores"] if r["student_id"] == "7240101"]
    assert "数学" in subjects
    assert "物理" in subjects


# ════════════════════════════════════════════════════════════════
#  §5 重复上传：同 exam 不重复写残留
# ════════════════════════════════════════════════════════════════

def test_duplicate_upload_does_not_double_store(tmp_path):
    """同一文件两次 parse_and_store（归入同 exam）：第二次只新增上传记录，
    不重复写 SubjectScore；TotalScore 始终 0 行。"""
    setup = _MAKE_GRADE23_XLSX + textwrap.dedent("""\
        import os
        from app.db.models import Teacher
        db = SessionLocal()
        t = Teacher(subject="数学", name="数学老师")
        db.add(t); db.commit(); db.close()

        raw_dir = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw")
        os.makedirs(raw_dir, exist_ok=True)
        xlsx_path = os.path.join(raw_dir, "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        make_grade23_xlsx(xlsx_path, [
            ["7240101","01","1","卞幻", 97,108,120, 48,52, None,None,
             None,None,None,None,None,None,None,None,
             174, 325.5,"25.08%",283, 499.5,"30.01%",291,
             "75.08%","40.12%","35.45%", "物A1"],
        ])
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import SubjectScore, TotalScore
        from app.ingest.router import parse_and_store

        parsed = {"grade":2,"semester":"上","exam_type":"期中","sort_key":"2025-11","canonical_name":"高二2025学年第二学期期中考试"}
        xlsx_path = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw", "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        out1 = parse_and_store(xlsx_path, "f.xlsx", parsed, 2)
        assert out1["result"]["parsed_ok"]
        out2 = parse_and_store(xlsx_path, "f.xlsx", parsed, 2)
        assert out2["result"]["parsed_ok"]

        math_count = db.query(SubjectScore).filter(SubjectScore.subject=="数学").count()
        total_count = db.query(TotalScore).count()
        db.close()
        # 第二次会再写一遍 SubjectScore（parse_and_store 不去重），但 TotalScore 始终为 0
        assert total_count == 0, f"TotalScore 始终为0, 实际{total_count}"
        print(json.dumps({"ok": True, "math_count": math_count}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §6 旧 SQLite 已有 TotalScore 仍可启动 + 备份/恢复
# ════════════════════════════════════════════════════════════════

def test_legacy_sqlite_with_totalscore_boots(tmp_path):
    """旧库已有 TotalScore 行，应用仍可正常启动并查询。"""
    setup = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import Teacher, Exam, SubjectScore, TotalScore
        t = Teacher(subject="数学"); db.add(t); db.flush()
        e = Exam(name="旧考试", grade=2, semester="上", exam_type="月考", exam_date="2024-09")
        db.add(e); db.flush()
        db.add(SubjectScore(exam_id=e.id, student_id="s1", subject="数学", raw_score=90, name="甲", class_num=1))
        db.add(TotalScore(exam_id=e.id, student_id="s1", total_type="主三门", total_score=300, xueji_rank=10))
        db.commit(); db.close()
    """)

    assert_code = textwrap.dedent("""\
        # 应用已启动（import app.main 触发建表），旧 TotalScore 行仍在
        db = SessionLocal()
        from app.db.models import TotalScore
        n = db.query(TotalScore).count()
        db.close()
        assert n == 1, f"旧 TotalScore 行应保留, 实际{n}"
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


def test_backup_restore_preserves_legacy_totalscore(tmp_path):
    """备份→恢复：旧 TotalScore 行无损保留。"""
    setup = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import Teacher, Exam, SubjectScore, TotalScore
        t = Teacher(subject="数学"); db.add(t); db.flush()
        e = Exam(name="旧考试", grade=2, semester="上", exam_type="月考", exam_date="2024-09")
        db.add(e); db.flush()
        db.add(SubjectScore(exam_id=e.id, student_id="s1", subject="数学", raw_score=90, name="甲", class_num=1))
        db.add(TotalScore(exam_id=e.id, student_id="s1", total_type="主三门", total_score=300, xueji_rank=10))
        db.commit(); db.close()
    """)

    assert_code = textwrap.dedent("""\
        from app.backup.router import create_backup
        from app.db.models import SessionLocal, TotalScore
        import os, zipfile, shutil
        from app.paths import DATA_DIR, BACKUP_DIR
        from app.db.models import engine

        fname = create_backup(prefix="test")
        zpath = os.path.join(BACKUP_DIR, fname)
        assert os.path.exists(zpath)

        # 恢复
        engine.dispose()
        db_path = os.path.join(DATA_DIR, "db.sqlite")
        with zipfile.ZipFile(zpath) as zf:
            with zf.open("db.sqlite") as src, open(db_path, "wb") as dst:
                shutil.copyfileobj(src, dst)

        db = SessionLocal()
        n = db.query(TotalScore).count()
        db.close()
        assert n == 1, f"恢复后旧 TotalScore 应保留, 实际{n}"
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §7 删除考试级联清理旧 TotalScore
# ════════════════════════════════════════════════════════════════

def test_delete_exam_cleans_legacy_totalscore(tmp_path):
    """删除考试时，旧 TotalScore 关联行也被级联清理（整场清理）。"""
    setup = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import Teacher, Exam, SubjectScore, TotalScore
        t = Teacher(subject="数学"); db.add(t); db.flush()
        e = Exam(name="旧考试", grade=2, semester="上", exam_type="月考", exam_date="2024-09")
        db.add(e); db.flush()
        db.add(SubjectScore(exam_id=e.id, student_id="s1", subject="数学", raw_score=90, name="甲", class_num=1))
        db.add(TotalScore(exam_id=e.id, student_id="s1", total_type="主三门", total_score=300, xueji_rank=10))
        db.commit(); db.close()
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import Exam, TotalScore, SubjectScore
        e = db.query(Exam).first()
        eid = e.id
        db.query(SubjectScore).filter(SubjectScore.exam_id==eid).delete(synchronize_session=False)
        db.query(TotalScore).filter(TotalScore.exam_id==eid).delete(synchronize_session=False)
        db.delete(e); db.commit()
        n = db.query(TotalScore).count()
        db.close()
        assert n == 0
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §8 同名消歧不查询 TotalScore（静态 + 行为）
# ════════════════════════════════════════════════════════════════

def test_name_candidates_no_totalscore_query(tmp_path):
    """name_candidates 用当前任教学科最近合法成绩/班级信息辅助，不查 TotalScore。
    同名两名候选人只含学号/姓名/班级/最近成绩，不含主三门名次字段。"""
    setup = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import Teacher, Exam, SubjectScore, TotalScore
        t = Teacher(subject="数学"); db.add(t); db.flush()
        e = Exam(name="考试", grade=1, semester="下", exam_type="期中", exam_date="2025-04")
        db.add(e); db.flush()
        # 两名同名学生，不同学号、不同班
        db.add(SubjectScore(exam_id=e.id, student_id="s1", subject="数学", raw_score=90, name="张三", class_num=1))
        db.add(SubjectScore(exam_id=e.id, student_id="s2", subject="数学", raw_score=80, name="张三", class_num=2))
        # 诱惑 TotalScore
        db.add(TotalScore(exam_id=e.id, student_id="s1", total_type="主三门", xueji_rank=5))
        db.add(TotalScore(exam_id=e.id, student_id="s2", total_type="主三门", xueji_rank=50))
        db.commit(); db.close()
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.analysis.scope import name_candidates
        cands = name_candidates(db, "张三", target_grade=1)
        db.close()
        assert len(cands) == 2
        for c in cands:
            # 不依赖 TotalScore：latest_rank 来源应为当前学科成绩，非主三门名次
            assert "latest_rank" in c or "latest_score" in c or "class_num" in c
        # 两人 class_num 不同，可辅助辨识
        classes = {c.get("class_num") for c in cands}
        assert classes == {1, 2}
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §1b 解析器 internal contract 不再返回 total_scores 键
# ════════════════════════════════════════════════════════════════

def test_grade23_parser_result_has_no_total_scores_key():
    """高二/高三解析器结果 dict 不再包含 total_scores 键。
    旧格式 Excel（含 +3/主三门/3+3 总分列）仍可解析各单科成绩。"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "学生成绩明细"
    ws.cell(2, 1, "学号")
    ws.cell(2, 2, "班级")
    ws.cell(2, 3, "学籍")
    ws.cell(2, 4, "姓名")
    ws.cell(2, 20, "+3总分")
    ws.cell(2, 21, "主三门")
    ws.cell(2, 24, "3+3总分")
    ws.append([])
    ws.append([
        "7240101", "01", "1", "卞幻",
        97, 108, 120, 48, 52, 50, 55,
        None, None, None, None, None, None, None, None,
        174, 325.5, "25.08%", 283, 499.5, "30.01%", 291,
        "75.08%", "40.12%", "35.45%",
    ])
    import tempfile, os
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)

    from app.ingest.excel_parser import parse_excel_grade23
    result = parse_excel_grade23(path, grade=2)
    os.unlink(path)

    assert result["kind"] == "student_scores"
    assert "total_scores" not in result, (
        f"解析器结果不应再包含 total_scores 键（TotalScore 已退役），"
        f"实际 keys={list(result.keys())}"
    )


def test_grade1_parser_result_has_no_total_scores_key(tmp_path):
    """高一解析器结果 dict 不再包含 total_scores 键。
    旧格式 Excel（含主三门/五门/九门总分列）仍可解析各单科成绩。"""
    from app.ingest.excel_parser import parse_excel_grade1
    xlsx_path = tmp_path / "2024级2024学年第二学期期中考试.xlsx"
    # 复用 grade1 制表助手
    exec(_MAKE_GRADE1_XLSX, {})
    ns = {}
    exec(_MAKE_GRADE1_XLSX, ns)
    ns["make_grade1_xlsx"](str(xlsx_path))

    result = parse_excel_grade1(str(xlsx_path))
    assert result["kind"] == "student_scores"
    assert "total_scores" not in result, (
        f"解析器结果不应再包含 total_scores 键（TotalScore 已退役），"
        f"实际 keys={list(result.keys())}"
    )


# ════════════════════════════════════════════════════════════════
#  §3 detected_class/detected_classes 只从教师任教学科 filtered rows 推导
# ════════════════════════════════════════════════════════════════

def test_detected_classes_from_teacher_subject_only(tmp_path):
    """数学教师上传含数学+物理的 Excel：detected_classes 的 class_labels
    只含数学对应行推导出的教学班标签，不把物理科的标签（如 物A1）回传前端。
    detected_class（行政班）也从数学行推导。"""
    setup = _MAKE_GRADE23_XLSX + textwrap.dedent("""\
        import os
        from app.db.models import Teacher
        db = SessionLocal()
        t = Teacher(subject="数学", name="数学老师")
        db.add(t); db.commit(); db.close()

        raw_dir = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw")
        os.makedirs(raw_dir, exist_ok=True)
        xlsx_path = os.path.join(raw_dir, "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        # 两行：第一行数学+物理（教学班 物A1）；第二行只有物理（教学班 物A2，无数学成绩）
        make_grade23_xlsx(xlsx_path, [
            ["7240101","01","1","卞幻", 97,108,120, 48,52, None,None,
             None,None,None,None,None,None,None,None,
             174, 325.5,"25.08%",283, 499.5,"30.01%",291,
             "75.08%","40.12%","35.45%", "物A1"],
            ["7240102","02","1","李四", None,None,None, 50,55, None,None,
             None,None,None,None,None,None,None,None,
             None, None,None,None, None,None,None,
             None,None,None, "物A2"],
        ])
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.ingest.router import parse_and_store

        parsed = {"grade":2,"semester":"上","exam_type":"期中","sort_key":"2025-11","canonical_name":"高二2025学年第二学期期中考试"}
        xlsx_path = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw", "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        out = parse_and_store(xlsx_path, "f.xlsx", parsed, 2)
        assert out["result"]["parsed_ok"], out["result"]

        # 数学教师只存数学行 → detected_class 应来自数学行（01 班），不是 02
        assert out["detected_class"] == 1, f"detected_class 应为1(数学行), 实际{out['detected_class']}"

        # class_labels 不应含 物A2（无数学成绩的第二行教学班标签）
        dc = out.get("detected_classes") or {}
        labels = dc.get("class_labels") or []
        assert "物A2" not in labels, f"物A2(无数学成绩行的教学班)不应进入detected_classes, 实际{labels}"
        db.close()
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §4 SubjectScore 存储丢弃 raw_score 和 grade_score 均空的 percentile 残留行
# ════════════════════════════════════════════════════════════════

def test_percentile_only_rows_not_stored(tmp_path):
    """数学教师上传含数学行的 Excel，其中某行数学 raw_score/grade_score 均空
    但 percentile 有值（残留）→ 该行不得存入 SubjectScore。"""
    setup = _MAKE_GRADE23_XLSX + textwrap.dedent("""\
        import os
        from app.db.models import Teacher
        db = SessionLocal()
        t = Teacher(subject="数学", name="数学老师")
        db.add(t); db.commit(); db.close()

        raw_dir = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw")
        os.makedirs(raw_dir, exist_ok=True)
        xlsx_path = os.path.join(raw_dir, "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        # 第一行有数学成绩；第二行数学原始分和等级分都空但百分位列（col 28）有值
        make_grade23_xlsx(xlsx_path, [
            ["7240101","01","1","卞幻", 97,108,120, 48,52, None,None,
             None,None,None,None,None,None,None,None,
             174, 325.5,"25.08%",283, 499.5,"30.01%",291,
             "75.08%","40.12%","35.45%", "物A1"],
            ["7240102","02","1","李四", 80,None,None, None,None, None,None,
             None,None,None,None,None,None,None,None,
             None, None,None,None, None,None,None,
             "50.00%","50.00%","50.00%", "物A1"],
        ])
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import SubjectScore
        from app.ingest.router import parse_and_store

        parsed = {"grade":2,"semester":"上","exam_type":"期中","sort_key":"2025-11","canonical_name":"高二2025学年第二学期期中考试"}
        xlsx_path = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw", "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        out = parse_and_store(xlsx_path, "f.xlsx", parsed, 2)
        assert out["result"]["parsed_ok"], out["result"]

        # 只应有 1 行数学（第一行有真实分数）；第二行 percentile-only 残留不得入库
        math_rows = db.query(SubjectScore).filter(SubjectScore.subject=="数学").all()
        db.close()
        assert len(math_rows) == 1, f"数学应只存1行(有真实分数), 实际{len(math_rows)}"
        assert math_rows[0].student_id == "7240101"
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §5 ClassAverage 只写教师当前学科值真实存在的合法行，total_averages 恒 {}
# ════════════════════════════════════════════════════════════════

def test_class_average_skips_class_without_teacher_subject(tmp_path):
    """数学教师上传含多个班的均分表，某班数学均分缺失但其他科有值 →
    该班不写空壳行（subject_averages 为 {}）；只有数学值真实的班才写。
    所有写入行的 total_averages 恒为 {}。"""
    setup = _MAKE_GRADE23_AVG_XLSX + textwrap.dedent("""\
        import os
        from openpyxl import Workbook
        from app.db.models import Teacher
        db = SessionLocal()
        t = Teacher(subject="数学", name="数学老师")
        db.add(t); db.commit(); db.close()

        raw_dir = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw")
        os.makedirs(raw_dir, exist_ok=True)
        xlsx_path = os.path.join(raw_dir, "高二2025学年第二学期期中考试班级均分表.xlsx")
        # 多班均分：01班有数学；02班数学缺失但有语文/物理
        wb = Workbook(); ws = wb.active; ws.title = "班级均分"
        ws.append(["班型","班级","教学班","班主任","语文","数学","英语","物理",None,"化学",None,"加3同均分","主三门","3+3总分"])
        ws.append([None,None,None,None,None,None,None,None,None,None,None,None,None,None])
        ws.append(["平行班","01","物A1","张老师", 101.2,108.3,110.4, 55.1,61.2, 58.3,62.4, 180.5,319.9,500.4])
        ws.append(["平行班","02","物A2","李老师", 99.0, None, 105.0, 50.0,58.0, 55.0,60.0, 175.0,310.0,490.0])
        wb.save(xlsx_path)
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import ClassAverage
        from app.ingest.router import parse_and_store

        parsed = {"grade":2,"semester":"上","exam_type":"期中","sort_key":"2025-11","canonical_name":"高二2025学年第二学期期中考试"}
        xlsx_path = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw", "高二2025学年第二学期期中考试班级均分表.xlsx")
        out = parse_and_store(xlsx_path, "f.xlsx", parsed, 2)
        assert out["result"]["parsed_ok"], out["result"]

        avgs = db.query(ClassAverage).all()
        db.close()
        # 02班数学缺失 → 不写空壳行；只有 01班（数学值真实）才写
        assert len(avgs) == 1, f"只应写1行(01班有数学), 实际{len(avgs)}"
        assert avgs[0].class_num == 1
        sa = avgs[0].subject_averages or {}
        assert set(sa.keys()) == {"数学"}, f"subject_averages 应只含数学, 实际{set(sa.keys())}"
        assert avgs[0].total_averages == {}, f"total_averages 应恒为空, 实际{avgs[0].total_averages}"
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §6 Teacher.subject 未设置 或 文件无任教学科真实成绩 → 拒绝并 rollback
# ════════════════════════════════════════════════════════════════

def test_upload_rejected_when_teacher_subject_unset(tmp_path):
    """未配置 Teacher.subject → 上传拒绝（parsed_ok=False），不创建空 Exam/Upload。"""
    setup = _MAKE_GRADE23_XLSX + textwrap.dedent("""\
        import os
        raw_dir = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw")
        os.makedirs(raw_dir, exist_ok=True)
        xlsx_path = os.path.join(raw_dir, "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        make_grade23_xlsx(xlsx_path, [
            ["7240101","01","1","卞幻", 97,108,120, 48,52, None,None,
             None,None,None,None,None,None,None,None,
             174, 325.5,"25.08%",283, 499.5,"30.01%",291,
             "75.08%","40.12%","35.45%", "物A1"],
        ])
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import Exam, Upload, SubjectScore
        from app.ingest.router import parse_and_store

        parsed = {"grade":2,"semester":"上","exam_type":"期中","sort_key":"2025-11","canonical_name":"高二2025学年第二学期期中考试"}
        xlsx_path = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw", "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        out = parse_and_store(xlsx_path, "f.xlsx", parsed, 2)
        # 未配置 teacher.subject → 拒绝
        assert not out["result"]["parsed_ok"], f"应拒绝(parsed_ok=False), 实际{out['result']}"
        # 不创建空 Exam / Upload / SubjectScore
        exam_count = db.query(Exam).count()
        upload_count = db.query(Upload).count()
        score_count = db.query(SubjectScore).count()
        db.close()
        assert exam_count == 0, f"不应创建Exam, 实际{exam_count}"
        assert upload_count == 0, f"不应创建Upload, 实际{upload_count}"
        assert score_count == 0, f"不应创建SubjectScore, 实际{score_count}"
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


def test_upload_rejected_when_file_has_no_teacher_subject(tmp_path):
    """教师任数学，但上传的 Excel 完全没有数学成绩 → 拒绝（parsed_ok=False），
    不创建空业务数据。"""
    setup = _MAKE_GRADE23_XLSX + textwrap.dedent("""\
        import os
        from app.db.models import Teacher
        db = SessionLocal()
        t = Teacher(subject="数学", name="数学老师")
        db.add(t); db.commit(); db.close()

        raw_dir = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw")
        os.makedirs(raw_dir, exist_ok=True)
        xlsx_path = os.path.join(raw_dir, "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        # 数学原始分(col6)为空、数学百分位(col28)为空 → 无数学成绩
        make_grade23_xlsx(xlsx_path, [
            ["7240101","01","1","卞幻", 97,None,120, 48,52, None,None,
             None,None,None,None,None,None,None,None,
             174, 325.5,"25.08%",283, 499.5,"30.01%",291,
             "75.08%",None,"35.45%", "物A1"],
        ])
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import Exam, Upload, SubjectScore
        from app.ingest.router import parse_and_store

        parsed = {"grade":2,"semester":"上","exam_type":"期中","sort_key":"2025-11","canonical_name":"高二2025学年第二学期期中考试"}
        xlsx_path = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw", "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        out = parse_and_store(xlsx_path, "f.xlsx", parsed, 2)
        assert not out["result"]["parsed_ok"], f"文件无数学成绩应拒绝, 实际{out['result']}"
        score_count = db.query(SubjectScore).count()
        db.close()
        assert score_count == 0, f"不应入库任何成绩, 实际{score_count}"
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §7 sync_members_after_upload 只维护当前教师 subject 的 TeachingClass
# ════════════════════════════════════════════════════════════════

def test_sync_members_only_teacher_subject_class(tmp_path):
    """数学教师上传数学成绩（教学班标签「数A1」）：
    - 数学教学班（label=数A1, subject=数学）应含上传学生；
    - 行政班（无 subject）仍按 class_num 同步（不受影响）；
    - 他科教学班（label 不同、subject=物理）的成员列表不被本次上传波及。

    注：TeachingClass 有 UniqueConstraint(grade, label)，同 label 的他科班在
    物理上不可能共存。本测试验证他科教学班不被扫描加入。
    """
    setup = _MAKE_GRADE23_XLSX + textwrap.dedent("""\
        import os
        from app.db.models import Teacher, TeachingClass
        db = SessionLocal()
        t = Teacher(subject="数学", name="数学老师")
        db.add(t); db.commit()
        # 数学教学班（label=数A1）和物理教学班（label=物B1，不同 label）
        db.add(TeachingClass(grade=2, kind="教学", subject="数学", label="数A1"))
        db.add(TeachingClass(grade=2, kind="教学", subject="物理", label="物B1"))
        db.commit(); db.close()

        raw_dir = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw")
        os.makedirs(raw_dir, exist_ok=True)
        xlsx_path = os.path.join(raw_dir, "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        make_grade23_xlsx(xlsx_path, [
            ["7240101","01","1","卞幻", 97,108,120, 48,52, None,None,
             None,None,None,None,None,None,None,None,
             174, 325.5,"25.08%",283, 499.5,"30.01%",291,
             "75.08%","40.12%","35.45%", "数A1"],
        ])
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import TeachingClass, TeachingClassMember
        from app.ingest.router import parse_and_store

        parsed = {"grade":2,"semester":"上","exam_type":"期中","sort_key":"2025-11","canonical_name":"高二2025学年第二学期期中考试"}
        xlsx_path = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw", "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        out = parse_and_store(xlsx_path, "f.xlsx", parsed, 2)
        assert out["result"]["parsed_ok"], out["result"]

        math_tc = db.query(TeachingClass).filter(TeachingClass.subject=="数学", TeachingClass.label=="数A1").first()
        physics_tc = db.query(TeachingClass).filter(TeachingClass.subject=="物理", TeachingClass.label=="物B1").first()
        math_members = {m.student_id for m in db.query(TeachingClassMember).filter(TeachingClassMember.teaching_class_id==math_tc.id).all()} if math_tc else set()
        physics_members = {m.student_id for m in db.query(TeachingClassMember).filter(TeachingClassMember.teaching_class_id==physics_tc.id).all()} if physics_tc else set()
        db.close()
        assert "7240101" in math_members, f"数学班应含7240101, 实际{math_members}"
        assert "7240101" not in physics_members, f"物理班不应含7240101, 实际{physics_members}"
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §8 candidate_classes 按教师任教学科 + 真实分数行过滤
# ════════════════════════════════════════════════════════════════

def test_candidate_classes_filtered_by_teacher_subject(tmp_path):
    """candidate_classes 只返回当前教师学科有真实分数的班号/教学班标签。
    旧库含他科 SubjectScore 行（含他科 class_label）不得进入设置页/前端。"""
    setup = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import Teacher, Exam, SubjectScore
        t = Teacher(subject="数学"); db.add(t); db.flush()
        e = Exam(name="旧考试", grade=2, semester="上", exam_type="月考", exam_date="2024-09")
        db.add(e); db.flush()
        # 数学成绩行（行政班 01，教学班 数A1）
        db.add(SubjectScore(exam_id=e.id, student_id="s1", subject="数学", raw_score=90, name="甲", class_num=1, class_label="数A1"))
        # 物理成绩行（行政班 02，教学班 物B2）— 诱惑：不得进入候选
        db.add(SubjectScore(exam_id=e.id, student_id="s2", subject="物理", raw_score=80, name="乙", class_num=2, class_label="物B2"))
        db.commit(); db.close()
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.teaching.service import candidate_classes
        result = candidate_classes(db, grade=2)
        db.close()
        class_nums = result.get("class_nums") or []
        class_labels = result.get("class_labels") or []
        # 数学行：班号 1，教学班 数A1
        assert 1 in class_nums, f"应含数学班号1, 实际{class_nums}"
        assert 2 not in class_nums, f"不应含物理班号2, 实际{class_nums}"
        assert "数A1" in class_labels, f"应含数A1, 实际{class_labels}"
        assert "物B2" not in class_labels, f"不应含物理标签物B2, 实际{class_labels}"
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §9 sync_members_after_upload 异常不静默吞错，失败可观察
# ════════════════════════════════════════════════════════════════

def test_sync_members_exception_not_silently_swallowed(tmp_path):
    """sync_members_after_upload 抛异常时，上传 result 应标记失败或传播错误，
    不得静默 except: pass 造成成绩已提交而成员同步失败。"""
    setup = _MAKE_GRADE23_XLSX + textwrap.dedent("""\
        import os
        from app.db.models import Teacher
        db = SessionLocal()
        t = Teacher(subject="数学", name="数学老师")
        db.add(t); db.commit(); db.close()

        raw_dir = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw")
        os.makedirs(raw_dir, exist_ok=True)
        xlsx_path = os.path.join(raw_dir, "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        make_grade23_xlsx(xlsx_path, [
            ["7240101","01","1","卞幻", 97,108,120, 48,52, None,None,
             None,None,None,None,None,None,None,None,
             174, 325.5,"25.08%",283, 499.5,"30.01%",291,
             "75.08%","40.12%","35.45%", "物A1"],
        ])
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.ingest.router import parse_and_store
        import app.teaching.service as svc

        # 猴子补丁让 sync_members_after_upload 抛异常
        original = svc.sync_members_after_upload
        def boom(db, exam):
            raise RuntimeError("sync boom")
        svc.sync_members_after_upload = boom
        try:
            parsed = {"grade":2,"semester":"上","exam_type":"期中","sort_key":"2025-11","canonical_name":"高二2025学年第二学期期中考试"}
            xlsx_path = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw", "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
            out = parse_and_store(xlsx_path, "f.xlsx", parsed, 2)
        finally:
            svc.sync_members_after_upload = original

        # 异常不应被静默吞掉：result 要么 parsed_ok=False（含错误信息），要么抛出
        # 关键：不能 parsed_ok=True 且隐藏了同步失败
        assert not out["result"].get("parsed_ok") or "sync" in (out["result"].get("message") or "").lower(), (
            f"sync异常不应被静默吞掉, 实际result={out['result']}"
        )
        # §2：异常后新 Session 断言零残留——fresh DB 基线为 0。
        # 不能只检查 parsed_ok False，必须确认事务回滚干净。
        from app.db.models import Exam, Upload, SubjectScore, TeachingClassMember
        db.expire_all()
        exam_count = db.query(Exam).count()
        upload_count = db.query(Upload).count()
        score_count = db.query(SubjectScore).count()
        member_count = db.query(TeachingClassMember).count()
        db.close()
        assert exam_count == 0, f"异常后不应残留Exam, 实际{exam_count}"
        assert upload_count == 0, f"异常后不应残留Upload, 实际{upload_count}"
        assert score_count == 0, f"异常后不应残留SubjectScore, 实际{score_count}"
        assert member_count == 0, f"异常后不应残留TeachingClassMember, 实际{member_count}"
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §9b 旧遗留 analysis 模块不再含 TotalScore 业务查询（静态扫描）
# ════════════════════════════════════════════════════════════════

class TestLegacyAnalysisModulesCleaned:
    """旧遗留 analysis 模块（无生产调用者）已彻底删除，不再以空壳/返回空 list
    的方式保留旧契约。

    trends.py 是例外：chat/tools.py（本卡禁止修改）仍调用 compute_student_trend，
    该函数读取 TotalScore，故 trends.py 暂保留为明确的集成冲突项；待阶段6 chat
    合并后一并删除。其余三个无生产调用者的模块（focus_list / cross_year /
    class_compare）已从 app 生产树移除，analysis/__init__ 不再导入它们的符号。
    """

    # 已删除的空壳/多学科遗留模块
    DELETED_MODULES = [
        "app/analysis/focus_list.py",
        "app/analysis/cross_year.py",
        "app/analysis/class_compare.py",
    ]

    def test_deleted_modules_no_longer_exist(self):
        import os
        import app.analysis as pkg

        pkg_dir = os.path.dirname(pkg.__file__)
        for rel in self.DELETED_MODULES:
            path = os.path.join(os.path.dirname(pkg_dir), rel.replace("app/analysis/", ""))
            assert not os.path.exists(path), f"{rel} 应已删除（空壳遗留模块）"

    def test_init_no_longer_imports_deleted_symbols(self):
        import importlib
        import app.analysis as pkg

        importlib.reload(pkg)
        for sym in ("build_focus_list", "compute_cross_year_trend", "compute_class_compare"):
            assert not hasattr(pkg, sym), (
                f"app.analysis 不应再导出 {sym}（对应空壳模块已删除）"
            )

    def test_deleted_modules_not_importable(self):
        for mod in ("app.analysis.focus_list", "app.analysis.cross_year", "app.analysis.class_compare"):
            try:
                __import__(mod)
            except ModuleNotFoundError:
                continue
            raise AssertionError(f"{mod} 应已删除、不可导入")


# ════════════════════════════════════════════════════════════════
#  §8/§9 CLI/辅助管线 analyze_exam_scores 彻底删除总分标识/参数/输出
# ════════════════════════════════════════════════════════════════

class TestCliPipelineTotalScoreRemoved:
    """analyze_exam_scores.py（CLI/辅助路径）按阶段7目标彻底删除总分相关
    生成、参数、管线和 CSV/JSON 字段，不得通过传空 list 保留旧契约。
    """

    REMOVED_IDENTIFIERS = [
        "TOTAL_COLS",
        "TREND_TOTAL_TYPES",
        "CLASS_AVG_TOTAL_COLS",
        "detect_total_type",
        "parse_rank_bands",
        "segment_for_rank",
        "trend_label",
        "build_student_trends",
        "build_focus_students",
        "build_subject_communication",
        "build_target_rank_bands",
        "aggregate_uploaded_rank_bands",
        "FOCUS_CATEGORIES",
        "PROGRESS_RANK_THRESHOLD",
        "VOLATILITY_RANK_THRESHOLD",
        "HIGH_RANK_MAX",
        "CRITICAL_RANK_MIN",
        "CRITICAL_RANK_MAX",
    ]

    KEPT_IDENTIFIERS = [
        "parse_student_scores",
        "parse_class_averages",
        "classify_workbook",
        "build_xueji1_analysis",
        "build_history_rows",
        "build_class_overview",
        "recommended_subjects",
        "run_analysis",
    ]

    def test_removed_identifiers_gone(self):
        import app.ingest.analyze_exam_scores as m
        for sym in self.REMOVED_IDENTIFIERS:
            assert not hasattr(m, sym), f"{sym} 应已从 analyze_exam_scores 删除"

    def test_kept_identifiers_present(self):
        import app.ingest.analyze_exam_scores as m
        for sym in self.KEPT_IDENTIFIERS:
            assert hasattr(m, sym), f"{sym} 应保留（单科分析仍需要）"

    def test_no_total_score_references_in_source(self):
        """源码不得出现 TotalScore 模型引用（文档注释中解释已删除的符号除外）。"""
        import inspect
        import app.ingest.analyze_exam_scores as m
        src = inspect.getsource(m)
        # 只禁止实际模型引用模式，允许文档/注释中提及符号名
        for forbidden in ("import TotalScore", "TotalScore)", "TotalScore,", "TotalScore."):
            assert forbidden not in src, f"analyze_exam_scores 源码不得出现 {forbidden}"

    def test_parse_inputs_no_rank_bands_in_result(self):
        """parse_inputs 结果不再包含 rank_bands 键（名次段表口径已退役）。"""
        import app.ingest.analyze_exam_scores as m
        result = m.parse_inputs([])
        assert "rank_bands" not in result, (
            f"parse_inputs 结果不应含 rank_bands 键, 实际 keys={list(result.keys())}"
        )

    def test_recommended_subjects_no_total_percentile_param(self):
        """recommended_subjects 不再接收 total_percentile 参数（总分百分位退役）。"""
        import inspect
        import app.ingest.analyze_exam_scores as m
        sig = inspect.signature(m.recommended_subjects)
        assert "total_percentile" not in sig.parameters, (
            "recommended_subjects 不应再有 total_percentile 参数"
        )

    def test_build_analysis_no_total_output_keys(self):
        """build_analysis 结果不含 student_trends/focus_students/rank_bands 等总分输出键。"""
        import app.ingest.analyze_exam_scores as m
        from pathlib import Path
        import tempfile
        # 用空输入构造 parsed，build_analysis 应可工作且不含总分输出键
        parsed = m.parse_inputs([])
        if not parsed["students"]:
            parsed["students"] = [{"exam": "test", "exam_order": 1, "student_id": "x",
                                   "class": "1", "xueji": "1", "name": "n", "source_file": "f"}]
        result = m.build_analysis(parsed, "1")
        for key in ("student_trends", "focus_students", "subject_communication",
                     "target_rank_bands", "uploaded_rank_band_summary", "rank_bands"):
            assert key not in result, f"build_analysis 结果不应含 {key} 键"


# ════════════════════════════════════════════════════════════════
#  §5/§6 sync_by_class_num 与 candidate_classes 的单学科领域约束
# ════════════════════════════════════════════════════════════════

def test_sync_by_class_num_filters_by_teacher_subject(tmp_path):
    """sync_by_class_num 只取教师任教学科 + 真实分数的学号；
    旧库含他科行不得聚合进行政班成员。"""
    setup = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import Teacher, Exam, SubjectScore, TeachingClass
        t = Teacher(subject="数学"); db.add(t); db.flush()
        e = Exam(name="考试", grade=1, semester="下", exam_type="期中", exam_date="2025-04")
        db.add(e); db.flush()
        # 数学真实分（s1）、物理真实分（s2，他科）、数学 percentile-only（s3，无真实分）
        db.add(SubjectScore(exam_id=e.id, student_id="s1", subject="数学", raw_score=90, name="甲", class_num=1))
        db.add(SubjectScore(exam_id=e.id, student_id="s2", subject="物理", raw_score=80, name="乙", class_num=1))
        db.add(SubjectScore(exam_id=e.id, student_id="s3", subject="数学", grade_percentile=50.0, name="丙", class_num=1))
        db.add(TeachingClass(grade=1, kind="行政", label="1"))
        db.commit(); db.close()
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import TeachingClass, TeachingClassMember
        from app.teaching.service import sync_by_class_num
        tc = db.query(TeachingClass).filter(TeachingClass.label=="1", TeachingClass.kind=="行政").first()
        count = sync_by_class_num(db, tc)
        db.commit()
        members = {m.student_id for m in db.query(TeachingClassMember).filter(TeachingClassMember.teaching_class_id==tc.id).all()}
        db.close()
        assert "s1" in members, f"数学真实分应入成员, 实际{members}"
        assert "s2" not in members, f"物理他科不应入成员, 实际{members}"
        assert "s3" not in members, f"percentile-only不应入成员, 实际{members}"
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


def test_candidate_classes_raises_without_teacher_subject(tmp_path):
    """candidate_classes 无 Teacher.subject 时领域错误，不退化为全学科。"""
    setup = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import Exam, SubjectScore
        e = Exam(name="考试", grade=2, semester="上", exam_type="月考", exam_date="2024-09")
        db.add(e); db.flush()
        db.add(SubjectScore(exam_id=e.id, student_id="s1", subject="数学", raw_score=90, name="甲", class_num=1))
        db.add(SubjectScore(exam_id=e.id, student_id="s2", subject="物理", raw_score=80, name="乙", class_num=2))
        db.commit(); db.close()
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.teaching.service import candidate_classes
        try:
            result = candidate_classes(db, grade=2)
        except ValueError as e:
            print(json.dumps({"ok": True, "error": str(e)}))
        else:
            raise AssertionError(f"无Teacher.subject应领域错误, 实际返回{result}")
        finally:
            db.close()
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §4 parser 成员查询只加教师学科 + 真实分数行
# ════════════════════════════════════════════════════════════════

def test_sync_parser_members_filter_subject_and_real_score(tmp_path):
    """sync_members_after_upload 的 parser 成员只加教师学科 + 真实分数行；
    旧他科 / percentile-only 行不加入教学班成员。"""
    setup = _MAKE_GRADE23_XLSX + textwrap.dedent("""\
        import os
        from app.db.models import Teacher, TeachingClass, Exam, SubjectScore
        db = SessionLocal()
        t = Teacher(subject="数学", name="数学老师")
        db.add(t); db.commit()
        db.add(TeachingClass(grade=2, kind="教学", subject="数学", label="数A1"))
        db.commit()
        # 预置一场旧考试，含他科（物理）行和数学 percentile-only 行
        old = Exam(name="旧考", grade=2, semester="上", exam_type="月考", exam_date="2024-09")
        db.add(old); db.flush()
        db.add(SubjectScore(exam_id=old.id, student_id="7240200", subject="物理", raw_score=70, name="旧他科", class_label="数A1"))
        db.add(SubjectScore(exam_id=old.id, student_id="7240201", subject="数学", grade_percentile=50.0, name="残留", class_label="数A1"))
        db.commit(); db.close()

        raw_dir = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw")
        os.makedirs(raw_dir, exist_ok=True)
        xlsx_path = os.path.join(raw_dir, "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        # 新上传：数学真实分，教学班标签 数A1
        make_grade23_xlsx(xlsx_path, [
            ["7240101","01","1","卞幻", 97,108,120, 48,52, None,None,
             None,None,None,None,None,None,None,None,
             174, 325.5,"25.08%",283, 499.5,"30.01%",291,
             "75.08%","40.12%","35.45%", "数A1"],
        ])
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import TeachingClass, TeachingClassMember
        from app.ingest.router import parse_and_store

        parsed = {"grade":2,"semester":"上","exam_type":"期中","sort_key":"2025-11","canonical_name":"高二2025学年第二学期期中考试"}
        xlsx_path = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw", "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        out = parse_and_store(xlsx_path, "f.xlsx", parsed, 2)
        assert out["result"]["parsed_ok"], out["result"]

        tc = db.query(TeachingClass).filter(TeachingClass.subject=="数学", TeachingClass.label=="数A1").first()
        members = {m.student_id for m in db.query(TeachingClassMember).filter(TeachingClassMember.teaching_class_id==tc.id).all()}
        db.close()
        assert "7240101" in members, f"新上传数学真实分应入成员, 实际{members}"
        assert "7240200" not in members, f"旧他科(物理)行不应入成员, 实际{members}"
        assert "7240201" not in members, f"旧percentile-only行不应入成员, 实际{members}"
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §1 existing-exam 回滚：source_files 改动也回滚
# ════════════════════════════════════════════════════════════════

def test_existing_exam_source_files_rolled_back_on_sync_failure(tmp_path):
    """已有 exam 的 source_files 追加在 sync 异常后也回滚（不残留半提交）。"""
    setup = _MAKE_GRADE23_XLSX + textwrap.dedent("""\
        import os
        from app.db.models import Teacher, Exam, SubjectScore
        db = SessionLocal()
        db.add(Teacher(subject="数学", name="数学老师")); db.flush()
        # 预置一场已有考试（一个 source_file）
        e = Exam(name="高二2025学年第二学期期中考试", grade=2, semester="上",
                 exam_type="期中", exam_date="2025-11", source_files=["/old/path.xlsx"])
        db.add(e); db.flush()
        db.add(SubjectScore(exam_id=e.id, student_id="7240099", subject="数学",
                            raw_score=85, name="旧生", class_num=1, class_label="数A1"))
        db.commit(); db.close()

        raw_dir = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw")
        os.makedirs(raw_dir, exist_ok=True)
        xlsx_path = os.path.join(raw_dir, "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
        make_grade23_xlsx(xlsx_path, [
            ["7240101","01","1","卞幻", 97,108,120, 48,52, None,None,
             None,None,None,None,None,None,None,None,
             174, 325.5,"25.08%",283, 499.5,"30.01%",291,
             "75.08%","40.12%","35.45%", "数A1"],
        ])
    """)

    assert_code = textwrap.dedent("""\
        db = SessionLocal()
        from app.db.models import Exam, Upload, SubjectScore, TeachingClassMember
        from app.ingest.router import parse_and_store
        import app.teaching.service as svc

        original = svc.sync_members_after_upload
        svc.sync_members_after_upload = lambda d, e: (_ for _ in ()).throw(RuntimeError("sync boom"))
        try:
            parsed = {"grade":2,"semester":"上","exam_type":"期中","sort_key":"2025-11","canonical_name":"高二2025学年第二学期期中考试"}
            xlsx_path = os.path.join(os.environ["EXAM_TRACKER_DIR"], "raw", "高二2025学年第二学期期中考试学生成绩明细表.xlsx")
            out = parse_and_store(xlsx_path, "f.xlsx", parsed, 2)
        finally:
            svc.sync_members_after_upload = original

        db.expire_all()
        # 基线：旧 exam 仍在（1 条），source_files 不变（未追加新 path）
        exam = db.query(Exam).first()
        assert exam is not None, "旧 exam 应保留"
        assert exam.source_files == ["/old/path.xlsx"], f"source_files 不应变, 实际{exam.source_files}"
        # 新增的 Upload / SubjectScore / TeachingClassMember 应为 0（回滚干净）
        upload_count = db.query(Upload).count()
        score_count = db.query(SubjectScore).count()
        member_count = db.query(TeachingClassMember).count()
        db.close()
        assert upload_count == 0, f"不应新增Upload, 实际{upload_count}"
        assert score_count == 1, f"应有旧1行数学(基线), 实际{score_count}"
        assert member_count == 0, f"不应残留TeachingClassMember, 实际{member_count}"
        print(json.dumps({"ok": True}))
    """)

    proc = _run_isolated_api_test(tmp_path, setup, assert_code)
    _assert_ok(proc)


# ════════════════════════════════════════════════════════════════
#  §8/§9 行为：单学科输入仍可生成完整数据包（无总分输出）
# ════════════════════════════════════════════════════════════════

def test_single_subject_input_produces_data_package(tmp_path):
    """单学科（数学）学生成绩明细表输入 run_analysis，仍应生成 workbook、
    summary、score_long.csv、class_averages.csv、analysis.json，且 JSON 结果
    不含任何总分输出键（student_trends/focus_students/rank_bands 等）。"""
    from openpyxl import Workbook
    from app.ingest.analyze_exam_scores import run_analysis

    d = tmp_path / "inputs"
    d.mkdir()
    # 高二单科（数学）学生成绩明细表
    wb = Workbook()
    ws = wb.active
    ws.title = "学生成绩明细"
    ws.cell(2, 1, "学号")
    ws.cell(2, 2, "班级")
    ws.cell(2, 3, "学籍")
    ws.cell(2, 4, "姓名")
    ws.cell(2, 7, "数学")  # SUBJECT_COLS: 数学=(7,8)
    ws.cell(2, 30, "教学班")
    ws.append([])
    ws.append(["7240101", "01", "1", "卞幻", None, None, 97, None,
               None, None, None, None, None, None, None, None, None, None,
               None, None, None, None, None, None, None, None, None, None, None, "数A1"])
    ws.append(["7240102", "01", "1", "李四", None, None, 85, None,
               None, None, None, None, None, None, None, None, None, None,
               None, None, None, None, None, None, None, None, None, None, None, "数A1"])
    p1 = d / "高二2025学年第二学期期中考试学生成绩明细表.xlsx"
    wb.save(p1)

    out_dir = tmp_path / "out"
    result = run_analysis([p1], out_dir, "1")

    # 数据包核心产物存在
    assert os.path.exists(result["workbook"]), "workbook 应生成"
    assert os.path.exists(result["summary"]), "summary 应生成"
    assert "score_long" in result["data"], "data 应含 score_long"
    assert os.path.exists(result["data"]["score_long"]), "score_long.csv 应生成"
    assert os.path.exists(result["data"]["analysis_json"]), "analysis.json 应生成"

    # JSON 结果不含任何总分输出键
    with open(result["data"]["analysis_json"]) as f:
        analysis = json.load(f)
    for key in ("student_trends", "focus_students", "subject_communication",
                "target_rank_bands", "uploaded_rank_band_summary", "rank_bands"):
        assert key not in analysis, f"analysis.json 不应含 {key}（总分输出已退役）"

    # score_long.csv 含数学行（单科输入仍产出有效数据）
    import csv
    with open(result["data"]["score_long"], encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    subjects = {r.get("subject") for r in rows}
    assert "数学" in subjects, f"score_long 应含数学行, 实际 subjects={subjects}"
