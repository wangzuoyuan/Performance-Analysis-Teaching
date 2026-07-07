import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.db.models import (
    Base, ClassRoster, HomeworkRecord, HomeworkSemester, SpecialRecord,
    TeachingClass, TeachingClassMember,
)
from app.homework.parser import parse_name_action
from app.homework.service import add_semester, dashboard, student_summary, warnings


def make_db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine)()


def seed_scope(db):
    db.add_all([
        ClassRoster(student_id="A1", name="同名", excluded=0),
        ClassRoster(student_id="A2", name="同名", excluded=0),
        ClassRoster(student_id="B1", name="乙", excluded=0),
        ClassRoster(student_id="X1", name="排除", excluded=1),
        TeachingClass(id=1, grade=2, label="物A1", kind="教学", sort_order=1),
        TeachingClass(id=2, grade=2, label="物A2", kind="教学", sort_order=2),
        TeachingClassMember(teaching_class_id=1, student_id="A1"),
        TeachingClassMember(teaching_class_id=1, student_id="A2"),
        TeachingClassMember(teaching_class_id=1, student_id="X1"),
        TeachingClassMember(teaching_class_id=2, student_id="A1"),
        TeachingClassMember(teaching_class_id=2, student_id="B1"),
        HomeworkSemester(
            id=1, name="测试学期", start_date="2026-03-01",
            end_date="2026-07-01", is_current=1,
        ),
    ])
    db.commit()


def record(db, sid, day, status="缺交", subject="物理", evaluation=None):
    db.add(HomeworkRecord(
        student_id=sid, date=day, subject=subject,
        submission_status=status, evaluation=evaluation,
    ))


def test_all_scope_is_configured_union_and_deduplicated():
    db = make_db()
    seed_scope(db)
    record(db, "A1", "2026-03-01")
    record(db, "B1", "2026-03-01")
    record(db, "X1", "2026-03-01")
    db.commit()
    result = dashboard(db, "2026-03-01", "2026-03-31")
    assert result["scope"]["member_count"] == 3
    assert result["kpi"]["total_misses"] == 2
    assert {x["student_id"] for x in result["rankings"]["missing"]} == {"A1", "B1"}
    # 全勤只发给「所在班当期确有收交记录」且零缺交的学生
    assert {x["student_id"] for x in result["honors"]["full_attendance"]} == {"A2"}


def test_consecutive_two_yellow_three_red_and_submission_breaks():
    db = make_db()
    seed_scope(db)
    for day in ("2026-03-02", "2026-03-03"):
        record(db, "A1", day)
    for day in ("2026-03-01", "2026-03-02", "2026-03-03"):
        record(db, "B1", day)
    db.commit()
    result = warnings(db, "2026-03-01", "2026-03-31")
    assert {x["student_id"] for x in result["warning"]} == {"A1"}
    assert {x["student_id"] for x in result["serious"]} == {"B1"}

    record(db, "B1", "2026-03-04", status="已交")
    db.commit()
    result = warnings(db, "2026-03-01", "2026-03-31")
    assert "B1" not in {x["student_id"] for x in result["serious"]}


def test_leave_excluded_and_dashboard_awards():
    db = make_db()
    seed_scope(db)
    record(db, "A1", "2026-03-01")
    db.add(SpecialRecord(student_id="A1", date="2026-03-01", type="请假"))
    for day, evaluation in (
        ("2026-06-20", "优秀"), ("2026-06-21", "认真"), ("2026-06-22", "进步"),
    ):
        record(db, "A2", day, status="已交", evaluation=evaluation)
    record(db, "A1", "2026-06-20", status="已交", evaluation="马虎")
    record(db, "A1", "2026-06-21", status="已交", evaluation="潦草")
    for day in ("2026-03-02", "2026-03-03", "2026-03-04"):
        db.add(SpecialRecord(student_id="B1", date=day, type="忘带"))
    db.commit()
    result = dashboard(db, "2026-03-01", "2026-07-01")
    assert result["kpi"]["total_misses"] == 0
    assert {x["student_id"] for x in result["honors"]["excellent"]} == {"A2"}
    assert {x["student_id"] for x in result["warnings"]["forgot"]} == {"B1"}
    assert {x["student_id"] for x in result["warnings"]["quality"]} == {"A1"}
    assert result["submission_rates"]
    assert result["semester_compare"][0]["name"] == "测试学期"


def test_scope_falls_back_to_full_roster_without_classes():
    """老库升级后尚未配置教学班：回落全花名册，看板不能失明。"""
    db = make_db()
    db.add_all([
        ClassRoster(student_id="R1", name="甲", excluded=0),
        ClassRoster(student_id="R2", name="乙", excluded=0),
    ])
    db.commit()
    record(db, "R1", "2026-03-01")
    db.commit()
    result = dashboard(db, "2026-03-01", "2026-03-31")
    assert result["scope"]["member_count"] == 2
    assert result["kpi"]["total_misses"] == 1
    assert {x["student_id"] for x in result["rankings"]["missing"]} == {"R1"}
    assert {x["student_id"] for x in result["honors"]["full_attendance"]} == {"R2"}


def test_other_students_submission_does_not_break_streak():
    """别的学生一条「已交」不能终结本人的连续缺交预警。"""
    db = make_db()
    seed_scope(db)
    for day in ("2026-03-01", "2026-03-02", "2026-03-03"):
        record(db, "B1", day)
    record(db, "A1", "2026-03-04", status="已交", evaluation="优秀")
    db.commit()
    result = warnings(db, "2026-03-01", "2026-03-31")
    assert "B1" in {x["student_id"] for x in result["serious"]}


def test_excellent_follows_selected_range():
    """优秀榜跟随所选区间，不锚定「今天-30 天」（历史区间也有效）。"""
    db = make_db()
    seed_scope(db)
    for day, evaluation in (
        ("2026-03-02", "优秀"), ("2026-03-03", "认真"), ("2026-03-04", "进步"),
    ):
        record(db, "A2", day, status="已交", evaluation=evaluation)
    db.commit()
    result = dashboard(db, "2026-03-01", "2026-03-31")
    assert {x["student_id"] for x in result["honors"]["excellent"]} == {"A2"}
    assert result["rankings"]["excellent"][0]["student_id"] == "A2"


def test_semester_compare_excludes_leave_days():
    """学期对比与 KPI 同口径：请假当天的缺交不计入。"""
    db = make_db()
    seed_scope(db)
    record(db, "A1", "2026-03-05")
    db.add(SpecialRecord(student_id="A1", date="2026-03-05", type="请假"))
    db.commit()
    result = dashboard(db, "2026-03-01", "2026-03-31")
    assert result["semester_compare"][0]["misses"] == 0


def test_add_semester_duplicate_raises_value_error():
    db = make_db()
    seed_scope(db)
    add_semester(db, "上学期", "2025-09-01", "2026-01-31")
    with pytest.raises(ValueError):
        add_semester(db, "上学期", "2025-09-01", "2026-01-31")


def test_student_summary_includes_excluded_student():
    """指定具体学生查询时不排除 excluded（模块既有口径）。"""
    db = make_db()
    seed_scope(db)
    record(db, "X1", "2026-03-10")
    db.commit()
    result = student_summary(db, student_id="X1")
    assert result["total_misses"] == 1


def test_export_daily_report_excludes_submitted(tmp_path, monkeypatch):
    """「已交/优秀」记录不得混进当日缺交 Excel。"""
    from openpyxl import load_workbook

    from app.homework import export as export_mod

    monkeypatch.setattr(export_mod, "EXPORT_DIR", str(tmp_path))
    db = make_db()
    seed_scope(db)
    record(db, "B1", "2026-03-01")
    record(db, "A1", "2026-03-01", status="已交", evaluation="优秀")
    db.commit()
    path = export_mod.export_daily_report("2026-03-01", db=db)
    names = [row[1].value for row in load_workbook(path).active.iter_rows(min_row=2)]
    assert "乙" in names
    assert "同名" not in names


def test_name_action_parser_supports_status_evaluation_and_special():
    names = {"张三"}
    missing = parse_name_action("张三数学缺交订正", names)
    assert missing["submission_status"] == "缺交"
    assert missing["subject"] == "数学"
    excellent = parse_name_action("张三物理优秀", names)
    assert excellent["submission_status"] == "已交"
    assert excellent["evaluation"] == "优秀"
    forgot = parse_name_action("张三忘带英语作业", names)
    assert forgot["special_type"] == "忘带"


def seed_name_only_class(db):
    """一个只配了「仅姓名」占位成员的教学班（还没上传成绩、没建花名册）。"""
    db.add_all([
        TeachingClass(id=1, grade=2, label="物A1", kind="教学", sort_order=1),
        TeachingClassMember(
            teaching_class_id=1, student_id="_anon:方石", name="方石", source="manual"
        ),
        TeachingClassMember(
            teaching_class_id=1, student_id="_anon:顾陈旗", name="顾陈旗", source="manual"
        ),
        HomeworkSemester(
            id=1, name="测试学期", start_date="2026-03-01",
            end_date="2026-07-01", is_current=1,
        ),
    ])
    db.commit()


def test_name_only_members_counted_as_valid_students():
    """回归：仅姓名成员曾因 members_of 排除 _anon: 而使教学班「0 名有效学生」。

    启动迁移会给仅姓名成员补花名册行，之后作业看板即把他们算作有效学生。"""
    from app.db.migrate_teaching import _backfill_anon_member_roster

    db = make_db()
    seed_name_only_class(db)
    created = _backfill_anon_member_roster(db)
    db.commit()
    assert created == 2
    result = dashboard(db, "2026-03-01", "2026-03-31", teaching_class_id=1)
    assert result["scope"]["member_count"] == 2


def test_smart_input_records_name_only_member(monkeypatch):
    """回归：为仅姓名成员录缺交曾提示「未匹配到当前教学班学生」。

    直连内存库跑通 hw_smart_input 端点：应匹配到姓名、补建花名册行、写入缺交，
    并在看板 KPI 里可见。"""
    import asyncio
    from app.homework import router as hw_router

    db = make_db()
    seed_name_only_class(db)
    monkeypatch.setattr(hw_router, "get_db", lambda: iter([db]))
    monkeypatch.setattr(hw_router, "export_daily_report", lambda *a, **k: None)

    payload = hw_router.SmartInputPayload(
        raw_text="方石\n顾陈旗物理缺交",
        teaching_class_id=1,
        date="2026-03-05",
        confirm=True,
    )
    resp = asyncio.get_event_loop().run_until_complete(
        hw_router.hw_smart_input(payload)
    )
    assert resp["success"] is True
    assert resp["added_count"] == 2

    # 两名仅姓名学生都补建了花名册行，缺交记录进入看板口径
    assert db.query(ClassRoster).filter(
        ClassRoster.student_id == "_anon:方石"
    ).count() == 1
    result = dashboard(db, "2026-03-01", "2026-03-31", teaching_class_id=1)
    assert result["kpi"]["total_misses"] == 2


def test_rekey_scopes_anon_ids_per_class_and_migrates_data():
    """迁移：旧「全局按姓名」占位学号按教学班隔离；单班占位的花名册/缺交一并改指。"""
    from app.db.migrate_teaching import _rekey_anon_members_class_scoped

    db = make_db()
    db.add_all([
        TeachingClass(id=1, grade=2, label="物A1", kind="教学", sort_order=1),
        TeachingClass(id=2, grade=2, label="物B3", kind="教学", sort_order=2),
        # 两个班各有一个同名「王某」（旧格式共用一个占位学号）
        TeachingClassMember(teaching_class_id=1, student_id="_anon:王某", name="王某"),
        TeachingClassMember(teaching_class_id=2, student_id="_anon:王某", name="王某"),
        # 只有一个班有的「独苗」，带花名册和一条缺交
        TeachingClassMember(teaching_class_id=1, student_id="_anon:独苗", name="独苗"),
        ClassRoster(student_id="_anon:独苗", name="独苗", excluded=0),
        HomeworkRecord(
            student_id="_anon:独苗", date="2026-03-01", subject="物理",
            submission_status="缺交",
        ),
    ])
    db.commit()

    changed = _rekey_anon_members_class_scoped(db)
    db.commit()
    assert changed == 3  # 两个王某 + 一个独苗

    ids = {m.teaching_class_id: m.student_id for m in db.query(TeachingClassMember)
           .filter(TeachingClassMember.name == "王某").all()}
    assert ids == {1: "_anon:1:王某", 2: "_anon:2:王某"}

    # 独苗只属一个班：花名册与缺交都改指到新学号
    assert db.query(ClassRoster).filter(ClassRoster.student_id == "_anon:1:独苗").count() == 1
    assert db.query(ClassRoster).filter(ClassRoster.student_id == "_anon:独苗").count() == 0
    assert db.query(HomeworkRecord).filter(
        HomeworkRecord.student_id == "_anon:1:独苗"
    ).count() == 1

    # 幂等：再跑一次不再改动
    assert _rekey_anon_members_class_scoped(db) == 0


def test_same_name_miss_does_not_leak_across_classes(monkeypatch):
    """端到端：两个班同名仅姓名学生，给 A 班录缺交不应出现在 B 班看板。"""
    import asyncio
    from app.homework import router as hw_router
    from app.teaching.service import anon_sid_for

    db = make_db()
    db.add_all([
        TeachingClass(id=1, grade=2, label="物A1", kind="教学", sort_order=1),
        TeachingClass(id=2, grade=2, label="物B3", kind="教学", sort_order=2),
        TeachingClassMember(teaching_class_id=1, student_id=anon_sid_for("王某", 1), name="王某"),
        TeachingClassMember(teaching_class_id=2, student_id=anon_sid_for("王某", 2), name="王某"),
        HomeworkSemester(
            id=1, name="测试学期", start_date="2026-03-01",
            end_date="2026-07-01", is_current=1,
        ),
    ])
    db.commit()

    monkeypatch.setattr(hw_router, "get_db", lambda: iter([db]))
    monkeypatch.setattr(hw_router, "export_daily_report", lambda *a, **k: None)
    payload = hw_router.SmartInputPayload(
        raw_text="王某物理缺交", teaching_class_id=1, date="2026-03-05", confirm=True,
    )
    resp = asyncio.get_event_loop().run_until_complete(hw_router.hw_smart_input(payload))
    assert resp["success"] is True

    a = dashboard(db, "2026-03-01", "2026-03-31", teaching_class_id=1)
    b = dashboard(db, "2026-03-01", "2026-03-31", teaching_class_id=2)
    assert a["kpi"]["total_misses"] == 1
    assert b["kpi"]["total_misses"] == 0
