from pathlib import Path

from openpyxl import Workbook

from app.ingest.excel_parser import parse_excel_grade23
from app.ingest.excel_parser import parse_excel_grade1_student_scores


def make_grade23_student_workbook(path: Path):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "学生成绩明细"
    ws.cell(2, 1, "学号")
    ws.cell(2, 2, "班级")
    ws.cell(2, 3, "学籍")
    ws.cell(2, 4, "姓名")
    ws.cell(2, 20, "+3总分")
    ws.cell(2, 21, "主三门")
    ws.cell(2, 24, "3+3总分")

    # 高二/高三学生成绩表是固定列位解析：5-7 为语数英原始分，
    # 8-19 为选考科目原始/等级分，20-26 为总分/百分位/排名，27-29 为语数英百分位。
    ws.append([])
    ws.append([
        "7240101", "01", "1", "卞幻",
        97, 108, 120,
        48, 52,
        50, 55,
        None, None,
        None, None,
        None, None,
        None, None,
        174,
        325.5, "25.08%", 283,
        499.5, "30.01%", 291,
        "75.08%", "40.12%", "35.45%",
    ])
    ws.append([
        "7240102", "01", "1", "学生乙",
        90, 100, 110,
        None, None,
        45, 50,
        None, None,
        None, None,
        None, None,
        None, None,
        95,
        300, "40.00%", 350,
        395, "45.00%", 420,
        "60.00%", "55.00%", "50.00%",
    ])
    workbook.save(path)


def test_parse_grade23_sample_totals(tmp_path):
    """阶段7：total_scores 已退役——解析器不再返回该键。
    旧格式 Excel（含 +3/主三门/3+3 总分列）仍可正常解析各单科成绩。"""
    sample = tmp_path / "高二2025学年第二学期期中考试学生成绩明细表.xlsx"
    make_grade23_student_workbook(sample)
    result = parse_excel_grade23(str(sample), grade=2)

    assert result["kind"] == "student_scores"
    assert len(result["students"]) == 2
    # total_scores 键不再存在（TotalScore 退役）
    assert "total_scores" not in result
    # 各单科成绩仍可正常解析
    subjects = {
        row["subject"]: row
        for row in result["subject_scores"]
        if row["student_id"] == "7240101"
    }
    assert subjects["语文"]["raw_score"] == 97


def test_parse_grade23_sample_subjects(tmp_path):
    sample = tmp_path / "高二2025学年第二学期期中考试学生成绩明细表.xlsx"
    make_grade23_student_workbook(sample)
    result = parse_excel_grade23(str(sample), grade=2)

    subjects = {
        row["subject"]: row
        for row in result["subject_scores"]
        if row["student_id"] == "7240101"
    }
    assert subjects["语文"]["raw_score"] == 97
    assert subjects["语文"]["grade_percentile"] == 0.7508
    assert subjects["物理"]["raw_score"] == 48
    assert subjects["物理"]["grade_score"] == 52
    assert "生物" not in subjects


def test_parse_grade1_percentiles(tmp_path):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "学生成绩明细"
    ws.append(["学生成绩（在籍）"])
    ws.append(["学号", "班级", "学籍", "姓名", "语文", None, "数学", None, "主三门", None, None, None])
    ws.append([None, None, None, None, "分数", "年级百分位", "分数", "年级百分位", "总分", "年级百分位", "学籍排名", "年级排名"])
    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:C3")
    ws.merge_cells("D2:D3")
    ws.merge_cells("E2:F2")
    ws.merge_cells("G2:H2")
    ws.merge_cells("I2:L2")
    ws.append(["7240101", "01", "1", "卞幻", 92, "45.01%", 106, "66.84%", 306, "47.88%", 283, 283])
    path = tmp_path / "2024级2024学年第二学期期中考试.xlsx"
    workbook.save(path)

    result = parse_excel_grade1_student_scores(path)

    # 阶段7：total_scores 已退役——解析器不再返回该键
    assert "total_scores" not in result
    subjects = {
        row["subject"]: row
        for row in result["subject_scores"]
        if row["student_id"] == "7240101"
    }
    assert subjects["语文"]["grade_percentile"] == 0.4501
    assert subjects["数学"]["grade_percentile"] == 0.6684


def test_parse_grade23_class_average_workbook(tmp_path):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "班级均分"
    ws.append(
        [
            "班型",
            "班级",
            "班主任",
            "语文",
            "数学",
            "英语",
            "物理",
            None,
            "化学",
            None,
            "加3同均分",
            "主三门",
            "3+3总分",
        ]
    )
    ws.append([None, None, None, None, None, None, "原始", "等级", "原始", "等级", None, None, None])
    ws.append(["平行班", "01", "张老师", 101.2, 108.3, 110.4, 55.1, 61.2, 58.3, 62.4, 180.5, 319.9, 500.4])
    ws.append([None, "02", "李老师", 99.1, 103.2, 109.3, 53.1, 59.2, 57.3, 60.4, 176.5, 311.6, 488.1])
    path = tmp_path / "高二2025学年第二学期期中考试班级均分表.xlsx"
    workbook.save(path)

    result = parse_excel_grade23(str(path), grade=2)

    assert result["kind"] == "class_averages"
    assert len(result["class_averages"]) == 2
    first = result["class_averages"][0]
    assert first["class_type"] == "平行班"
    assert first["class_num"] == 1
    assert first["teacher_name"] == "张老师"
    assert first["subject_averages"]["语文"] == 101.2
    assert first["subject_averages"]["物理_等级"] == 61.2
    # 阶段7：兼容读取含旧总分列的 Excel，但 parser 契约不得再生成/返回总分。
    assert "total_averages" not in first
    # 无教学班列时，subject_scores 的 class_label 缺省为 None（写库时回退 str(class_num)）
    assert first["class_label"] is None


def _grade23_workbook_with_teaching_class(path: Path):
    """带「教学班」列的高二学生成绩明细表（教学版新格式）。"""
    workbook = Workbook()
    ws = workbook.active
    ws.title = "学生成绩明细"
    ws.cell(2, 1, "学号")
    ws.cell(2, 2, "班级")
    ws.cell(2, 3, "学籍")
    ws.cell(2, 4, "姓名")
    ws.cell(2, 20, "+3总分")
    ws.cell(2, 21, "主三门")
    ws.cell(2, 24, "3+3总分")
    ws.cell(2, 30, "教学班")  # 新增的教学班列
    ws.append([])
    ws.append([
        "7240101", "01", "1", "卞幻",
        97, 108, 120, 48, 52, 50, 55,
        None, None, None, None, None, None, None, None,
        174, 325.5, "25.08%", 283, 499.5, "30.01%", 291,
        "75.08%", "40.12%", "35.45%",
        "物A1",  # col30 教学班
    ])
    ws.append([
        "7240102", "01", "1", "学生乙",
        90, 100, 110, None, None, 45, 50,
        None, None, None, None, None, None, None, None,
        95, 300, "40.00%", 350, 395, "45.00%", 420,
        "60.00%", "55.00%", "50.00%",
        "物A2",  # col30 教学班
    ])
    workbook.save(path)


def test_parse_grade23_with_teaching_class_column(tmp_path):
    sample = tmp_path / "高二2025学年第二学期期中考试学生成绩明细表.xlsx"
    _grade23_workbook_with_teaching_class(sample)
    result = parse_excel_grade23(str(sample), grade=2)

    assert result["kind"] == "student_scores"
    by_student = {}
    for row in result["subject_scores"]:
        by_student.setdefault(row["student_id"], []).append(row)
    assert by_student["7240101"][0]["class_label"] == "物A1"
    assert by_student["7240102"][0]["class_label"] == "物A2"
    # 学生层也带 class_label（供 detected_classes）
    students = {s["student_id"]: s for s in result["students"]}
    assert students["7240101"]["class_label"] == "物A1"


def test_parse_grade23_class_average_with_teaching_class(tmp_path):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "班级均分"
    ws.append(["班型", "班级", "教学班", "班主任", "语文", "数学", "英语", "加3同均分", "主三门", "3+3总分"])
    ws.append([None, None, None, None, None, None, None, None, None, None])
    ws.append(["平行班", "01", "物A1", "张老师", 101.2, 108.3, 110.4, 180.5, 319.9, 500.4])
    ws.append([None, "02", "物A2", "李老师", 99.1, 103.2, 109.3, 176.5, 311.6, 488.1])
    path = tmp_path / "高二2025学年第二学期期中考试班级均分表.xlsx"
    workbook.save(path)

    result = parse_excel_grade23(str(path), grade=2)
    assert result["kind"] == "class_averages"
    by_num = {c["class_num"]: c for c in result["class_averages"]}
    assert by_num[1]["class_label"] == "物A1"
    assert by_num[2]["class_label"] == "物A2"
